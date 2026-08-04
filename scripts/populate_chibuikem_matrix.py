import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TARGET_XLSX = 'C:/Users/IT_COMMS/Downloads/Chibuikem_Performance_Tracker.xlsx'

def main():
    # Define the 15 core modules data with start dates, shortened explanatory descriptions, and target dates formatted as dd/mm/yyyy
    core_modules = [
        {
            'name': 'Leave',
            'start_date': '06/01/2026',
            'desc': 'Manages time-off requests. It checks balances, tracks reliever assignments, and routes applications through manager and HR approval workflows.',
            'status': 'In Progress',
            'timeline': '06/07/2026',
            'challenges': 'Requires manual backfilling of historical leave records and comprehensive workflow testing to resolve bugs.'
        },
        {
            'name': 'Project',
            'start_date': '22/06/2026',
            'desc': 'Tracks all company projects and progress, cataloging completed, active, and future projects in a centralized dashboard.',
            'status': 'Not Started',
            'timeline': '14/07/2026',
            'challenges': 'Requires continuous sessions with the HCS to fine-tune the module, with AI assistance (Claude) needed for implementation.'
        },
        {
            'name': 'Finance & Payments',
            'start_date': '12/01/2026',
            'desc': 'Oversees financial operations. It tracks client invoices, logs vendor bills through approvals, maps salary entries, and breaks down payments by department.',
            'status': 'In Progress',
            'timeline': '18/07/2026',
            'challenges': 'Requires a collaborative session with the Accounts Lead to align on payment workflows and improvements.'
        },
        {
            'name': 'Performance Management (PMS)',
            'start_date': '19/03/2026',
            'desc': 'Coordinates appraisals and performance tracking. It manages goals, KPI targets, competency frameworks, core values evaluations, and CBTs.',
            'status': 'In Progress',
            'timeline': '22/08/2026',
            'challenges': 'Requires extensive validation starting with Computer-Based Training (CBT), with AI integration assistance planned for this complex and critical component.'
        },
        {
            'name': 'Tasks',
            'start_date': '04/11/2025',
            'desc': 'Primary task collaboration board. It allows creating, assigning, and tracking tasks for employees and departments with escalation rules for admins.',
            'status': 'In Progress',
            'timeline': '22/08/2026',
            'challenges': 'Requires testing and user feedback to optimize task board workflows.'
        },
        {
            'name': 'Help Desk',
            'start_date': '28/02/2026',
            'desc': 'IT and support ticketing system. It logs issues, assigns tickets to responsible teams, tracks ticket status, and monitors SLA deadlines.',
            'status': 'In Progress',
            'timeline': '22/08/2026',
            'challenges': 'Requires validation testing and user input to refine support ticket workflows.'
        },
        {
            'name': 'Attendance',
            'start_date': '02/07/2026',
            'desc': (
                "Central time-tracking hub. It logs check-ins/outs, enforces boundary validation via geo-fencing "
                "and IP restrictions, manages remote approvals, and keeps historical records.\n\n"
                "📌 [RECENT TASKS COMPLETED - 02/07/2026]:\n"
                "• Attendance Configuration: Implemented Admin feedback regarding tracking requirements and immediately implemented the changes in the ERP.\n"
                "• Export Optimization: Optimized the ERP Excel sheet export functionality to match the Admin department's required format.\n"
                "• Attendance Rules: Added 'early' status to replace 'present' for easier understanding.\n"
                "• Credit System: Updated the credit system.\n"
                "• Daily Report Mailer: Developed a daily report mailer that automatically emails the attendance report to HR at 11:00 AM and 11:00 PM every workday."
            ),
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires further alignment sessions with HR and HCS for system fine-tuning.'
        },
        {
            'name': 'Documentation',
            'start_date': '04/11/2025',
            'desc': 'Central company policy and SOP library. It provides secure access to guidelines, manual documents, and department-specific files.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires active user adoption monitoring and feedback to optimize system accessibility and performance.'
        },
        {
            'name': 'Reports & General Meeting',
            'start_date': '02/07/2026',
            'desc': (
                "Handles meeting outputs and administrative reports. It compiles weekly reports, catalogs Knowledge Sharing Session (KSS) files, and tracks meeting action items.\n\n"
                "📌 [RECENT TASKS COMPLETED - 02/07/2026]:\n"
                "• Teams Meeting Records Route: Added records route to automatically fetch transcript and attendance data from Teams meetings for the week, store them in the ERP, and autosend the compilation to the Admin department."
            ),
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': None
        },
        {
            'name': 'Feedback',
            'start_date': '30/10/2025',
            'desc': 'Secure employee feedback channel. It routes named or anonymous suggestions and concerns to administrators for categorization, assignment, and response.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires user testing to validate the anonymity and reliability of the feedback submission and response loops.'
        },
        {
            'name': 'Assets',
            'start_date': '04/11/2025',
            'desc': 'Registers and tracks physical assets. It manages equipment inventory, assigns items to employees, logs damage reports, and tracks maintenance and repairs.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': None
        },
        {
            'name': 'Correspondence',
            'start_date': '01/07/2026',
            'desc': (
                "Standardizes official documents. It processes reference/introduction letter requests, "
                "populates templates with profile data, and routes them for admin sign-off.\n\n"
                "📌 [RECENT TASKS COMPLETED - 01/07/2026]:\n"
                "• Reference Generator: Updated the reference generator to make references unique by year and sender (instead of recipient)."
            ),
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': None
        },
        {
            'name': 'Communications',
            'start_date': '25/02/2026',
            'desc': 'Coordinates announcements and meeting schedules. It handles targeted broadcasts, schedules meetings, sends invites, and automates reminders.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': None
        },
        {
            'name': 'Employees & Directory',
            'start_date': '19/01/2026',
            'desc': 'Maintains the system-wide user registry. It stores contact profiles, maps department headcount structures, and manages user roles and permissions.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': None
        },
        {
            'name': 'Shared Resources',
            'start_date': '12/03/2026',
            'desc': 'Manages bookable company assets. It provides self-service booking for shared resources like fleet vehicles and meeting rooms with reservation calendar checks.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires testing and user validation to refine vehicle and meeting room booking logic.'
        },
        {
            'name': 'Network Monitoring',
            'start_date': '05/07/2026',
            'desc': 'Monitors and logs company network traffic. It tracks visited domains, resolves employee profiles, identifies device specifications (MAC address, vendor, OS, browser), logs bandwidth consumption (upload/download), and alerts admins to unrecognized devices and suspicious activity.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires fine-tuning classification filters to minimize false-positive review alerts from system background noise.'
        },
        {
            'name': 'Employee Onboarding',
            'start_date': '07/07/2026',
            'desc': 'Public onboarding form for new hires. It collects personal, residential, and job details, auto-generates company email previews, tracks draft progress in local storage, and securely submits applications for HR approval.',
            'status': 'Completed',
            'timeline': 'Done',
            'challenges': 'Requires configuring public-facing API routes with honeypot fields to block automated spam submissions.'
        }
    ]

    print(f"Loading destination workbook from {TARGET_XLSX}...")
    dest_wb = openpyxl.load_workbook(TARGET_XLSX)
    dest_ws = dest_wb['Matrix']
    
    # 1. Clean up old merged cells below row 1
    initial_merged = list(dest_ws.merged_cells.ranges)
    print(f"Initial merged ranges in Matrix sheet: {len(initial_merged)}")
    ranges_to_keep = [r for r in initial_merged if r.min_row < 2]
    dest_ws.merged_cells.ranges = ranges_to_keep
    print(f"Preserved merged ranges: {len(dest_ws.merged_cells.ranges)}")
    
    # 2. Clear rows from 2 to max_row
    max_row = dest_ws.max_row
    if max_row >= 2:
        print(f"Deleting old rows from 2 to {max_row}...")
        dest_ws.delete_rows(2, max_row - 1)
        
    # Styles
    font_regular = Font(name='Calibri', size=11, color='FF374151')
    font_title = Font(name='Calibri', size=11, bold=True, color='FF111827')
    
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
    
    # Status styles
    status_styles = {
        'Completed': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF166534'),
            'fill': PatternFill(fill_type='solid', fgColor='FFDCFCE7')
        },
        'In Progress': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF854D0E'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEF9C3')
        },
        'Not Started': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF991B1B'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEE2E2')
        }
    }
    
    thin_border = Border(
        left=Side(style='thin', color='FFD1D5DB'),
        right=Side(style='thin', color='FFD1D5DB'),
        top=Side(style='thin', color='FFD1D5DB'),
        bottom=Side(style='thin', color='FFD1D5DB')
    )
    
    alignments = [
        Alignment(horizontal='center', vertical='center'),  # Col A (Date)
        Alignment(horizontal='center', vertical='center'),  # Col B (Modules)
        Alignment(horizontal='left', vertical='center', wrap_text=True), # Col C (Task / Description)
        Alignment(horizontal='center', vertical='center'),  # Col D (Timeline)
        Alignment(horizontal='center', vertical='center'),  # Col E (Status)
        Alignment(horizontal='left', vertical='center', wrap_text=True)  # Col F (Key Challenges / Risks)
    ]
    
    current_row = 2
    for mod in core_modules:
        if mod['name'] == 'Attendance':
            dest_ws.row_dimensions[current_row].height = 160.0
        elif mod['name'] == 'Correspondence':
            dest_ws.row_dimensions[current_row].height = 90.0
        elif mod['name'] == 'Reports & General Meeting':
            dest_ws.row_dimensions[current_row].height = 110.0
        else:
            dest_ws.row_dimensions[current_row].height = 42.0
        
        row_values = [mod['start_date'], mod['name'], mod['desc'], mod['timeline'], mod['status'], mod.get('challenges')]
        
        for col in range(1, 7):
            cell = dest_ws.cell(row=current_row, column=col)
            cell.value = row_values[col - 1]
            cell.border = thin_border
            cell.alignment = alignments[col - 1]
            cell.fill = fill_white
            
            if col == 2:
                cell.font = font_title
            elif col == 5:
                # Status styling
                s_val = row_values[col - 1]
                if s_val in status_styles:
                    cell.font = status_styles[s_val]['font']
                    cell.fill = status_styles[s_val]['fill']
                else:
                    cell.font = font_regular
            else:
                cell.font = font_regular
                
        current_row += 1
        
    # Auto-adjust column widths slightly for aesthetics
    widths = {
        'A': 15,
        'B': 32,
        'C': 65,
        'D': 18,
        'E': 20,
        'F': 25
    }
    for col_letter, w in widths.items():
        dest_ws.column_dimensions[col_letter].width = w
        
    print(f"Saving destination workbook to {TARGET_XLSX}...")
    dest_wb.save(TARGET_XLSX)
    print("Success! Core modules successfully populated with shortened descriptions.")

if __name__ == '__main__':
    main()
