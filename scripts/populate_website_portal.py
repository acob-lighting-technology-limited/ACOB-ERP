import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TARGET_XLSX = 'C:/Users/IT_COMMS/Downloads/Chibuikem_Performance_Tracker.xlsx'

def main():
    # Define the 7 engineering focal areas for website portal
    web_modules = [
        {
            'start_date': '22/06/2026',
            'focal_area': 'Commercial & Industrial (C&I) Projects',
            'desc': 'Implement classification for isolated and interconnected residential, industrial installations, and residential solar systems in the project directory.',
            'timeline': '06/07/2026',
            'status': 'Not Started',
            'challenges': 'Requires gathering the project data (images and descriptions) to populate, followed by AI assistance (Claude) for implementation.'
        },
        {
            'start_date': '22/06/2026',
            'focal_area': 'Healthcare Electrification & Solarization',
            'desc': 'Configure project schemas and metrics to track healthcare electrification, tertiary hospital solarization, and primary healthcare facility solar installations. Add some Kaduna sites and its matrix to the system.',
            'timeline': '06/07/2026',
            'status': 'In Progress',
            'challenges': 'Requires gathering the project data (images and descriptions) to populate, followed by AI assistance (Claude) for implementation.'
        },
        {
            'start_date': '22/06/2026',
            'focal_area': 'PUE (Productive Use of Energy)',
            'desc': 'Integrate Productive Use of Energy (PUE) metrics and categorization for economic and agricultural solar applications.',
            'timeline': '06/07/2026',
            'status': 'Not Started',
            'challenges': 'Requires gathering the project data (images and descriptions) to populate, followed by AI assistance (Claude) for implementation.'
        },
        {
            'start_date': '22/06/2026',
            'focal_area': 'Irrigation Systems Integration',
            'desc': 'Integrate agricultural solar irrigation systems tracking and environmental/farming impact metrics into the project portfolio.',
            'timeline': '06/07/2026',
            'status': 'In Progress',
            'challenges': 'Requires gathering the project data (images and descriptions) to populate, followed by AI assistance (Claude) for implementation.'
        },
        {
            'start_date': '13/07/2025',
            'focal_area': 'User Interface & Layout Updates',
            'desc': 'Designed layout adjustments, styled components with Tailwind CSS/shadcn, optimized mobile-first responsiveness, and added Framer Motion animations.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        },
        {
            'start_date': '16/07/2025',
            'focal_area': 'Content Management System (CMS) Integration',
            'desc': 'Configured Sanity CMS schemas, developed GROQ query modules, and set up the server/browser Sanity clients for dynamic data fetching.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        },
        {
            'start_date': '23/07/2025',
            'focal_area': 'AI Support Integration (ACOBot)',
            'desc': 'Developed the AI customer chatbot widget, configured Vercel AI SDK routes, and set up Groq/OpenAI API integrations.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        },
        {
            'start_date': '24/07/2025',
            'focal_area': 'Forms & Email Services',
            'desc': 'Developed quotation, job application, and contact forms with Zod schema validation and Resend API automated email dispatch.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': 'The system was built but suspended due to the MD requesting it not be used for security reasons, preferring inquiries to be sent directly to WhatsApp.'
        },
        {
            'start_date': '15/11/2025',
            'focal_area': 'SEO & Performance Optimization',
            'desc': 'Configured structured SEO metadata, generated sitemaps, optimized images, and implemented service worker caching for offline PWA support.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        },
        {
            'start_date': '16/12/2025',
            'focal_area': 'Code Quality & Restructuring',
            'desc': 'Refactored the codebase, modularized types, centralized app constants, and created system architecture guides (ARCHITECTURE.md).',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        },
        {
            'start_date': '16/12/2025',
            'focal_area': 'System Deployment & Environment',
            'desc': 'Configured environment variables validation, set up Resend and Sanity project parameters, and deployed the Next.js application to Vercel.',
            'timeline': 'Done',
            'status': 'Completed',
            'challenges': None
        }
    ]

    print(f"Loading workbook from {TARGET_XLSX}...")
    wb = openpyxl.load_workbook(TARGET_XLSX)
    ws = wb['Web Portal']
    
    # 1. Clean up old merged cells below row 1
    initial_merged = list(ws.merged_cells.ranges)
    print(f"Initial merged ranges in Web Portal sheet: {len(initial_merged)}")
    ranges_to_keep = [r for r in initial_merged if r.min_row < 2]
    ws.merged_cells.ranges = ranges_to_keep
    print(f"Preserved merged ranges: {len(ws.merged_cells.ranges)}")
    
    # 2. Clear rows from 2 to max_row
    max_row = ws.max_row
    if max_row >= 2:
        print(f"Deleting old rows from 2 to {max_row}...")
        ws.delete_rows(2, max_row - 1)
        
    # Styles
    font_regular = Font(name='Calibri', size=11, color='FF374151')
    font_title = Font(name='Calibri', size=11, bold=True, color='FF111827')
    
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
    
    # Status styles
    status_styles = {
        'Completed': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF166534'),
            'fill': PatternFill(fill_type='solid', fgColor='FFDCFCE7')
        },
        'In Progress': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF854D0E'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEF9C3')
        },
        'Not Started': {
            'font': Font(name='Calibri', size=11, bold=True, color='FF991B1B'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEE2E2')
        }
    }
    
    thin_border = Border(
        left=Side(style='thin', color='FFD1D5DB'),
        right=Side(style='thin', color='FFD1D5DB'),
        top=Side(style='thin', color='FFD1D5DB'),
        bottom=Side(style='thin', color='FFD1D5DB')
    )
    
    alignments = [
        Alignment(horizontal='center', vertical='center'),  # Col A (Date)
        Alignment(horizontal='center', vertical='center'),  # Col B (Focal Area)
        Alignment(horizontal='left', vertical='center', wrap_text=True), # Col C (Task / Description)
        Alignment(horizontal='center', vertical='center'),  # Col D (Timeline)
        Alignment(horizontal='center', vertical='center'),  # Col E (Status)
        Alignment(horizontal='left', vertical='center', wrap_text=True)  # Col F (Key Challenges / Risks)
    ]
    
    current_row = 2
    for mod in web_modules:
        ws.row_dimensions[current_row].height = 42.0
        
        row_values = [mod['start_date'], mod['focal_area'], mod['desc'], mod['timeline'], mod['status'], mod['challenges']]
        
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.value = row_values[col - 1]
            cell.border = thin_border
            cell.alignment = alignments[col - 1]
            cell.fill = fill_white
            
            if col == 2:
                cell.font = font_title
            elif col == 5:
                # Status styling
                s_val = row_values[col - 1]
                if s_val in status_styles:
                    cell.font = status_styles[s_val]['font']
                    cell.fill = status_styles[s_val]['fill']
                else:
                    cell.font = font_regular
            else:
                cell.font = font_regular
                
        current_row += 1
        
    # Auto-adjust column widths slightly for aesthetics
    widths = {
        'A': 15,
        'B': 32,
        'C': 65,
        'D': 18,
        'E': 20,
        'F': 25
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
        
    print(f"Saving workbook to {TARGET_XLSX}...")
    wb.save(TARGET_XLSX)
    print("Success! Web Portal sheet successfully populated with 7 focal areas.")

if __name__ == '__main__':
    main()
