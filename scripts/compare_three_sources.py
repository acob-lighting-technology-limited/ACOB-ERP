import openpyxl
import datetime
import dotenv
import psycopg2
import re

env = dotenv.dotenv_values('.env.local')
db_url = env.get('SUPABASE_POSTGRES_URL_NON_POOLING') or env.get('SUPABASE_POSTGRES_URL')
conn = psycopg2.connect(db_url)
cur = conn.cursor()

# Get DB profiles
cur.execute('SELECT id, full_name, first_name, last_name FROM profiles;')
db_profiles = cur.fetchall()
prof_by_id = {p[0]: p[1] for p in db_profiles}

def normalize(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s)).strip().lower()

MANUAL_NAME_MAP = {
    'shirley': 'Shirley Jackreece',
    'shirley jackreece': 'Shirley Jackreece',
    'john': 'John Dangana',
    'john dangana': 'John Dangana',
    'susan eze': 'Susan Eze',
    'emmanuel': 'Emmanuel Ibanga',
    'emmanuel ibanga': 'Emmanuel Ibanga',
    'vanessa': 'Vanessa Lawrence-Ukaegbu',
    'vanessa lawrence': 'Vanessa Lawrence-Ukaegbu',
    'thomas': 'Thomas Olobo',
    'philip': 'Philip Yakubu',
    'busayo': 'Busayo Kadiri',
    'edward': 'Edward Atoshi',
    'edward benson': 'Benson Ademoye',
    'andrew': 'Andrew Inegbedion',
    'kennedy': 'Kennedy Odenigbo',
    'tansi': 'Tansi',
    'tansi o.': 'Tansi',
    'tansi or': 'Tansi',
    'tansi org': 'Tansi',
    'tansi orhorhomuke': 'Tansi',
    'dennis': 'Dennis Innug',
    'lawrence': 'Lawrence Adukwu',
    'favour': 'Favour Onuoha-Eke',
    'favour eke': 'Favour Onuoha-Eke',
    'eke favour': 'Favour Onuoha-Eke',
    'jessica': 'Jessica Egeonu',
    'somtochukwu jessica': 'Jessica Egeonu',
    'somto': 'Jessica Egeonu',
    'somtochkwu': 'Jessica Egeonu',
    'bawa': 'Alhamdu Bawa',
    'bawa alhamdu': 'Alhamdu Bawa',
    'bawa alhmdu': 'Alhamdu Bawa',
    'samila': 'Agya Samila',
    'elijah': 'Elijah Isah',
    'elija isah': 'Elijah Isah',
    'elijah isha': 'Elijah Isah',
    'tunde': 'Tunde Ajayi',
    'patrick': 'Patrick Prosper',
    'pattrick prosper': 'Patrick Prosper',
    'arc emmanuel': 'Emmanuel Chinedu',
    'tochucku': 'Tochukwu Nnadozie',
    'tochucku anothny': 'Tochukwu Nnadozie',
    'tochuwku anothny': 'Tochukwu Nnadozie',
    'tochukwu anthony': 'Tochukwu Nnadozie',
    'samad': 'Abdulsamad Danmusa',
    'samad danmusa': 'Abdulsamad Danmusa',
    'danmusa abdul': 'Abdulsamad Danmusa',
    'abiodun': 'Sefui Abiodun',
    'chibuikim': 'Chibuikem Ilonze',
    'chibuikim ilonze': 'Chibuikem Ilonze',
    'ilonze chibuikem': 'Chibuikem Ilonze',
    'victor': 'Victor Onyeka',
    'usman': 'Usman Haruna',
    'umar garba': 'Umar Garba',
    'suraj': 'Surajo Idris',
    'suraj idris': 'Surajo Idris',
    'surajo idris': 'Surajo Idris',
    'caleb': 'Caleb Obiechina',
    'caleb obechina': 'Caleb Obiechina',
    'daniel': 'Daniel Osa-Egonwa',
    'daniel osa-egonwa': 'Daniel Osa-Egonwa',
    'anointing': 'Anointed Emoghene',
    'anointed emoghene': 'Anointed Emoghene',
    'oghenerune': 'Oghenerune Orhorhomuke',
    'benson': 'Benson Ademoye',
    'sommy': 'Sommy',
    'ayoola': 'Peter Ayoola',
    'ayoola peter': 'Peter Ayoola',
    'peter ayoola': 'Peter Ayoola',
    'onyeka': 'Onyeka Eze',
    'oyneka': 'Onyeka Eze',
    'onyekachukwu atishie': 'Onyekachukwu Atishie',
    'khleo': 'Kleopatra Aiyede',
    'khleo aiyede': 'Kleopatra Aiyede',
    'cleopatra': 'Kleopatra Aiyede',
    'abdul': 'Abdulmalik Abdulkarim',
    'haron': 'Ode Haron',
    'peace': 'Peace Dogara',
    'terna peace': 'Peace Terna',
    'peace terna': 'Peace Terna',
    'new daniel': 'Istifanus Daniel',
    'seun': 'Oluwaseun Awotona',
    'tobi': 'Oluwatobi Oladele',
    'oladele oluwatobi': 'Oluwatobi Oladele',
    'jerome peter': 'Jerome Egemasi',
    'egemasi jerome': 'Jerome Egemasi',
    'paul ogboji': 'Paul Ojonugua',
    'tomi bayode': 'Oluwatomi Bayode',
    'tomi': 'Oluwatomi Bayode',
    'afeez azeez': 'Azeez Afeez',
    'ishaku martins': 'Martins Ishaku',
    'lawrence aduku': 'Lawrence Adukwu',
    'lawrence aduwku': 'Lawrence Adukwu',
    'orisakwe joseph': 'Joseph Orisakwe',
    'peter taiwo': 'Taiwo Peter',
    'simon aseya': 'Aseya Simon',
    'susan charels': 'Susan Eze'
}

def find_profile(name):
    norm = normalize(name)
    if not norm or norm in ['total', 'names', 'date', 's/n']: return None
    if norm in MANUAL_NAME_MAP:
        mapped_name = MANUAL_NAME_MAP[norm]
        for p in db_profiles:
            if normalize(p[1]) == normalize(mapped_name):
                return p
    for p in db_profiles:
        full = normalize(p[1])
        fn = normalize(p[2])
        ln = normalize(p[3])
        if norm == full or norm == f"{fn} {ln}" or norm == f"{ln} {fn}":
            return p
    for p in db_profiles:
        if norm in normalize(p[1]):
            return p
    return None

MONTH_NAMES_MAP = {
    'JAN': 1, 'JANUARY': 1,
    'FEB': 2, 'FEBRUARY': 2,
    'MAR': 3, 'MARCH': 3,
    'APR': 4, 'APRIL': 4,
    'MAY': 5,
    'JUN': 6, 'JUNE': 6,
    'JUL': 7, 'JULY': 7
}

def parse_date_file2(c_val, sheet_name, title_str):
    target_month = None
    title_upper = (sheet_name + " " + title_str).upper()
    for m_str, m_num in MONTH_NAMES_MAP.items():
        if m_str in title_upper:
            target_month = m_num
            break
            
    dt = None
    if isinstance(c_val, (datetime.datetime, datetime.date)):
        dt = c_val.date() if isinstance(c_val, datetime.datetime) else c_val
    else:
        val_str = str(c_val or "").strip()
        m = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})', val_str)
        if m:
            d, m_val, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100: y += 2000
            # If m_val is target month, d is day!
            if target_month and m_val == target_month:
                dt = datetime.date(y, m_val, d)
            elif target_month and d == target_month:
                dt = datetime.date(y, d, m_val)
            else:
                try:
                    dt = datetime.date(y, m_val, d)
                except ValueError:
                    try:
                        dt = datetime.date(y, d, m_val)
                    except ValueError:
                        pass

    if dt and target_month and dt.month != target_month:
        try:
            dt = datetime.date(dt.year, dt.day, dt.month)
        except ValueError:
            dt = datetime.date(2026, target_month, dt.day)

    return dt

# --- PARSE FILE B: "1 Updated 2026 STAFF LUNCH RECORD (1).xlsx" ---
wb_new = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\1 Updated 2026 STAFF LUNCH RECORD (1).xlsx', data_only=True)

new_file_entries = {}
unmapped_names_new = set()

for s_name in wb_new.sheetnames:
    ws = wb_new[s_name]
    title = str(ws.cell(1, 1).value or "")
    
    # Find header row (either row 3 or row 4)
    header_row_idx = 3
    for r_check in range(1, 6):
        c2_val = str(ws.cell(r_check, 2).value or "").strip().upper()
        if 'NAMES' in c2_val:
            header_row_idx = r_check
            break
            
    date_cols = {}
    for c in range(3, ws.max_column + 1):
        c_val = ws.cell(header_row_idx, c).value
        if c_val and str(c_val).strip().upper() != 'TOTAL':
            dt = parse_date_file2(c_val, s_name, title)
            if dt:
                date_cols[c] = dt

    for r in range(header_row_idx + 1, ws.max_row + 1):
        name_val = ws.cell(r, 2).value
        if not name_val or str(name_val).strip().upper() in ['TOTAL', 'S/N', 'NAMES']:
            continue
        emp_raw = str(name_val).strip()
        prof = find_profile(emp_raw)
        if not prof:
            unmapped_names_new.add(emp_raw)
            continue
            
        uid = prof[0]
        full_name = prof[1]
        
        for c, dt in date_cols.items():
            val = ws.cell(r, c).value
            if isinstance(val, (int, float)) and val > 0:
                new_file_entries[(dt, uid)] = (full_name, val, s_name)

print(f"=== FILE 2: '1 Updated 2026 STAFF LUNCH RECORD (1).xlsx' ===")
print(f"Total mapped meal logs extracted: {len(new_file_entries)}")
print(f"Remaining unmapped names in File 2: {sorted(unmapped_names_new)}")

# --- PARSE FILE A: "ACOB PETTY CASH BOOKS.xlsx" ---
wb_old = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\ACOB PETTY CASH BOOKS.xlsx', data_only=True)
ws_lakita = wb_old['LAKITA FOOD 2026']
ws_ded = wb_old['FOOD DEDUCTION 2026']

headers_l = [ws_lakita.cell(1, c).value for c in range(1, ws_lakita.max_column + 1)]
headers_d = [ws_ded.cell(1, c).value for c in range(1, ws_ded.max_column + 1)]

emp_map_old = {}
for c in range(3, 51):
    h = headers_l[c-1] or headers_d[c-1]
    p = find_profile(h)
    if p:
        emp_map_old[c] = p

def parse_date_old(val):
    dt = None
    if isinstance(val, (datetime.datetime, datetime.date)): dt = val.date() if isinstance(val, datetime.datetime) else val
    elif isinstance(val, str):
        try: dt = datetime.datetime.strptime(val.strip(), '%Y-%m-%d').date()
        except: dt = None
    if dt:
        if dt.year == 2025: dt = datetime.date(2026, dt.month, dt.day)
        if dt == datetime.date(2026, 1, 31): dt = datetime.date(2026, 1, 30)
    return dt

old_file_entries = {}
for ws in [ws_lakita, ws_ded]:
    for r in range(1, ws.max_row + 1):
        dt = parse_date_old(ws.cell(r, 2).value)
        if dt:
            for c in range(3, 51):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v > 0:
                    if c in emp_map_old:
                        p = emp_map_old[c]
                        old_file_entries[(dt, p[0])] = (p[1], v)

print(f"\n=== FILE 1: 'ACOB PETTY CASH BOOKS.xlsx' ===")
print(f"Total mapped meal logs extracted: {len(old_file_entries)}")

# --- DB LOGS ---
cur.execute("SELECT date, user_id FROM attendance_lunch_log WHERE date >= '2026-01-01' AND date <= '2026-12-31';")
db_logs = set(cur.fetchall())
print(f"\n=== DATABASE: 'attendance_lunch_log' ===")
print(f"Total DB 2026 meal logs: {len(db_logs)}")

# --- MONTHLY SUMMARY TABLE ---
months = range(1, 8)
month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul']

print(f"\n=================================================================")
print(f"  3-WAY MONTHLY COMPARISON TABLE (Jan - Jul 2026)")
print(f"=================================================================\n")
print(f"{'Month':<8} | {'ACOB Petty Cash (File 1)':<24} | {'1 Updated Record (File 2)':<25} | {'Database (DB)':<13}")
print("-" * 75)

for m, m_name in zip(months, month_names):
    f1_cnt = len([k for k in old_file_entries.keys() if k[0].month == m])
    f2_cnt = len([k for k in new_file_entries.keys() if k[0].month == m])
    db_cnt = len([k for k in db_logs if k[0].month == m])
    print(f"{m_name:<8} | {f1_cnt:<24} | {f2_cnt:<25} | {db_cnt:<13}")

conn.close()
