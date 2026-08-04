import openpyxl
import datetime
import dotenv
import psycopg2
import re

env = dotenv.dotenv_values('.env.local')
db_url = env.get('SUPABASE_POSTGRES_URL_NON_POOLING') or env.get('SUPABASE_POSTGRES_URL')
conn = psycopg2.connect(db_url)
cur = conn.cursor()

# Get DB profiles
cur.execute('SELECT id, full_name, first_name, last_name, employee_number, employment_status FROM profiles;')
db_profiles = cur.fetchall()
prof_by_id = {p[0]: p for p in db_profiles}

def normalize(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s)).strip().lower()

# Known manual name aliases mapping Excel Header -> DB Full Name
MANUAL_NAME_MAP = {
    'shirley': 'Shirley Jackreece',
    'john': 'John Dangana',
    'susan eze': 'Susan Eze',
    'emmanuel': 'Emmanuel Ibanga',
    'vanessa': 'Vanessa Lawrence-Ukaegbu',
    'thomas': 'Thomas Olobo',
    'philip': 'Philip Yakubu',
    'busayo': 'Busayo Kadiri',
    'edward': 'Edward Atoshi',
    'andrew': 'Andrew Inegbedion',
    'kennedy': 'Kennedy Odenigbo',
    'tansi': 'Tansi', # Unmapped / Exited
    'dennis': 'Dennis Innug',
    'lawrence': 'Lawrence Adukwu',
    'favour': 'Favour Onuoha-Eke',
    'jessica': 'Jessica Egeonu',
    'bawa': 'Alhamdu Bawa',
    'samila': 'Agya Samila',
    'elijah': 'Elijah Isah',
    'tunde': 'Tunde Ajayi',
    'patrick': 'Patrick Prosper',
    'arc emmanuel': 'Emmanuel Chinedu',
    'tochucku': 'Tochukwu Nnadozie',
    'samad': 'Abdulsamad Danmusa',
    'abiodun': 'Sefui Abiodun',
    'chibuikim': 'Chibuikem Ilonze',
    'victor': 'Victor Onyeka',
    'usman': 'Usman Haruna',
    'habeeb': 'Habeeb',
    'habib': 'Habeeb',
    'suraj': 'Surajo Idris',
    'caleb': 'Caleb Obiechina',
    'daniel': 'Daniel Osa-Egonwa',
    'anointing': 'Anointed Emoghene',
    'oghenerune': 'Oghenerune Orhorhomuke',
    'rune': 'Oghenerune Orhorhomuke',
    'benson': 'Benson Ademoye',
    'sommy': 'Sommy',
    'ayoola': 'Peter Ayoola',
    'gregory': 'Gregory',
    'onyeka': 'Onyeka Eze',
    'oyneka': 'Onyeka Eze',
    'khleo': 'Kleopatra Aiyede',
    'cleopatra': 'Kleopatra Aiyede',
    'abdul': 'Abdulmalik Abdulkarim',
    'haron': 'Ode Haron',
    'peace': 'Peace Dogara',
    'new daniel': 'Istifanus Daniel',
    'seun': 'Oluwaseun Awotona',
    'tobi': 'Oluwatobi Oladele',
    'noble': 'Noble'
}

def find_profile_for_header(header):
    norm = normalize(header)
    if not norm or norm == 'total' or norm == 'names' or norm == 'date': 
        return None
    
    # Check manual map first
    if norm in MANUAL_NAME_MAP:
        mapped_name = MANUAL_NAME_MAP[norm]
        for p in db_profiles:
            if normalize(p[1]) == normalize(mapped_name):
                return p
                
    # Direct match on full_name
    for p in db_profiles:
        if normalize(p[1]) == norm: 
            return p
            
    # Match on first_name
    for p in db_profiles:
        if normalize(p[2]) == norm:
            return p
            
    # Match on last_name
    for p in db_profiles:
        if normalize(p[3]) == norm:
            return p

    return None

# Load Excel workbook
wb = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\ACOB PETTY CASH BOOKS.xlsx', data_only=True)
ws_lakita = wb['LAKITA FOOD 2026']
ws_ded = wb['FOOD DEDUCTION 2026']

headers_l = [ws_lakita.cell(1, c).value for c in range(1, ws_lakita.max_column + 1)]
headers_d = [ws_ded.cell(1, c).value for c in range(1, ws_ded.max_column + 1)]

emp_map = {}
for c in range(3, 51):
    h_l = str(headers_l[c-1]).strip() if c <= len(headers_l) and headers_l[c-1] is not None else ""
    h_d = str(headers_d[c-1]).strip() if c <= len(headers_d) and headers_d[c-1] is not None else ""
    
    h = h_l or h_d
    p = find_profile_for_header(h)
    if p:
        emp_map[c] = p

def parse_date_fixed(val):
    dt = None
    if isinstance(val, datetime.datetime):
        dt = val.date()
    elif isinstance(val, datetime.date):
        dt = val
    elif isinstance(val, str):
        val = val.strip()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
            try:
                dt = datetime.datetime.strptime(val, fmt).date()
                break
            except ValueError:
                pass
    if not dt:
        return None
        
    # APPLY USER FIX 1: Correct 2025 year typos in 2026 sheet to 2026
    if dt.year == 2025:
        dt = datetime.date(2026, dt.month, dt.day)
        
    # APPLY USER FIX 4: Jan 31 / Jan 30 typo -> 2026-01-30 (Friday)
    if dt == datetime.date(2026, 1, 31):
        dt = datetime.date(2026, 1, 30)
        
    return dt

def extract_sheet_entries(ws):
    entries = {}
    for r in range(1, ws.max_row + 1):
        c2 = ws.cell(r, 2).value
        dt = parse_date_fixed(c2)
        if dt:
            for c in range(3, 51):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v > 0:
                    entries[(dt, c)] = v
    return entries

lakita_entries = extract_sheet_entries(ws_lakita)
ded_entries = extract_sheet_entries(ws_ded)

# Build combined Excel set of (date, user_id)
excel_set = set()
for (dt, c) in set(lakita_entries.keys()).union(set(ded_entries.keys())):
    if c in emp_map:
        user_id = emp_map[c][0]
        excel_set.add((dt, user_id))

# Fetch Database 2026 lunch log entries
cur.execute("SELECT date, user_id FROM attendance_lunch_log WHERE date >= '2026-01-01' AND date <= '2026-12-31';")
db_logs = cur.fetchall()
db_set = set((l[0], l[1]) for l in db_logs)

print("=================================================================")
print("  RECONCILIATION AFTER APPLYING EXCEL FIXES (Year 2026 & Jan 30)")
print("=================================================================\n")

print(f"Total 2026 lunch log entries in Database: {len(db_set)}")
print(f"Total 2026 mapped lunch log entries in Excel (Corrected): {len(excel_set)}")

in_excel_not_db = excel_set - db_set
in_db_not_excel = db_set - excel_set

print(f"\nNet Discrepancies:")
print(f"  - Present in Excel but MISSING IN DATABASE: {len(in_excel_not_db)}")
print(f"  - Present in Database but MISSING IN EXCEL: {len(in_db_not_excel)}")

# Monthly summary
months = range(1, 13)
month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

print("\n--- UPDATED MONTHLY RECONCILIATION SUMMARY (Excel vs Database) ---")
print(f"{'Month':<8} | {'Excel Meals':<12} | {'DB Meals':<10} | {'Missing in DB':<15} | {'Extra in DB':<12}")
print("-" * 65)

for m, m_name in zip(months, month_names):
    ex_m = set(e for e in excel_set if e[0].month == m)
    db_m = set(d for d in db_set if d[0].month == m)
    missing_db = ex_m - db_m
    extra_db = db_m - ex_m
    if len(ex_m) > 0 or len(db_m) > 0:
        print(f"{m_name:<8} | {len(ex_m):<12} | {len(db_m):<10} | {len(missing_db):<15} | {len(extra_db):<12}")

conn.close()
