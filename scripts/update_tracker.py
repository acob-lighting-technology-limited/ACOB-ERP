import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\Chibuikem_Performance_Tracker.xlsx')
ws = wb['System Control']

# Clear rows 2 and 3 completely first
for row in range(2, 15):
    for col in range(1, 7):
        ws.cell(row=row, column=col).value = None

def style(cell, font_color, bg_color, bold=False, halign='center', wrap=False, valign='center'):
    cell.font = Font(name='Calibri', size=11, bold=bold, color=font_color)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)

def write_row(ws, row, sn, focal, task, timeline, status, challenges, status_bg, status_fc):
    data = [sn, focal, task, timeline, status, challenges]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if col == 1:
            style(cell, 'FF374151', 'FFFFFFFF', halign='center')
        elif col == 2:
            style(cell, 'FF111827', 'FFFFFFFF', bold=True, halign='center')
        elif col == 3:
            style(cell, 'FF374151', 'FFFFFFFF', halign='left', wrap=True)
        elif col == 4:
            style(cell, 'FF374151', 'FFFFFFFF', halign='center')
        elif col == 5:
            style(cell, status_fc, status_bg, bold=True, halign='center')
        elif col == 6:
            style(cell, 'FF374151', 'FFFFFFFF', halign='left', wrap=True)

# Row 2 - Network Administration (restored)
write_row(
    ws, 2,
    sn=1,
    focal='Network Administration',
    task='Monitoring internet links and local networks via MikroTik routers and FortiGate firewall consoles.',
    timeline=None,
    status=None,
    challenges=None,
    status_bg='FFFFFFFF',
    status_fc='FF374151'
)
ws.row_dimensions[2].height = 40

# Row 3 - Email Security
write_row(
    ws, 3,
    sn=2,
    focal='Email Security',
    task=(
        'businessgrowth@acoblighting.com had over 300 unread emails with no subject clogging the inbox. '
        'Investigated and found that outsiders were faking our company email address to send spam, '
        'and all the failed delivery alerts were landing back on us. '
        'Set up email signing so recipients can verify our emails are genuinely from us, '
        'and configured a policy to send any faked emails straight to spam going forward.'
    ),
    timeline='22/06/2026',
    status='Completed',
    challenges=(
        'Email signing was never configured, leaving the domain open to impersonation. '
        'The spam policy is currently set to send faked emails to spam — '
        'needs to be upgraded to full rejection in 2-3 weeks after confirming no real emails are affected.'
    ),
    status_bg='FFDCFCE7',
    status_fc='FF166534'
)
ws.row_dimensions[3].height = 90

wb.save(r'C:\Users\IT_COMMS\Downloads\Chibuikem_Performance_Tracker.xlsx')
print('Done')
