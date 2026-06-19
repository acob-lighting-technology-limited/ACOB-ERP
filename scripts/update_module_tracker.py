import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
WORKSPACE_XLSX = 'C:/Users/IT_COMMS/GitHubProjects/ACOB-Signature-Creator/ACOB_ERP_Module_Tracker.xlsx'
DOWNLOADS_XLSX = 'C:/Users/IT_COMMS/Downloads/ACOB_ERP_Module_Tracker.xlsx'

def main():
    print(f"Loading workbook from {WORKSPACE_XLSX}...")
    wb = openpyxl.load_workbook(WORKSPACE_XLSX)
    ws = wb['Module Tracker']
    
    # 1. Clean up old merged cells below row 4
    # We keep only merged cells in rows 1-2 (Title and Subtitle)
    initial_merged = list(ws.merged_cells.ranges)
    print(f"Initial merged ranges: {len(initial_merged)}")
    ranges_to_keep = [r for r in initial_merged if r.min_row <= 2]
    ws.merged_cells.ranges = ranges_to_keep
    print(f"Preserved merged ranges: {len(ws.merged_cells.ranges)}")
    
    # 2. Clear rows from 5 to max_row
    max_row = ws.max_row
    if max_row >= 5:
        print(f"Deleting old rows from 5 to {max_row}...")
        ws.delete_rows(5, max_row - 4)
    
    # Define styling helpers
    font_regular = Font(name='Arial', size=9, color='FF374151')
    font_title = Font(name='Arial', size=9, bold=True, color='FF111827')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    
    fill_white = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
    fill_header = PatternFill(fill_type='solid', fgColor='FF1A3D28')
    
    # Status styles
    status_styles = {
        'Completed': {
            'font': Font(name='Arial', size=9, bold=True, color='FF166534'),
            'fill': PatternFill(fill_type='solid', fgColor='FFDCFCE7')
        },
        'In Progress': {
            'font': Font(name='Arial', size=9, bold=True, color='FF854D0E'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEF9C3')
        },
        'Not Started': {
            'font': Font(name='Arial', size=9, bold=True, color='FF991B1B'),
            'fill': PatternFill(fill_type='solid', fgColor='FFFEE2E2')
        }
    }
    
    thin_border = Border(
        left=Side(style='thin', color='FFD1D5DB'),
        right=Side(style='thin', color='FFD1D5DB'),
        top=Side(style='thin', color='FFD1D5DB'),
        bottom=Side(style='thin', color='FFD1D5DB')
    )
    
    # Alignments
    alignments = [
        Alignment(horizontal='center', vertical='center'),  # Col A (Index)
        Alignment(horizontal='left', vertical='center'),    # Col B (Section)
        Alignment(horizontal='left', vertical='center'),    # Col C (Title)
        Alignment(horizontal='left', vertical='center'),    # Col D (URL Path)
        Alignment(horizontal='center', vertical='center'),  # Col E (Location)
        Alignment(horizontal='center', vertical='center'),  # Col F (Status)
        Alignment(horizontal='left', vertical='center', wrap_text=True), # Col G (Description)
        Alignment(horizontal='center', vertical='center')   # Col H (Timeline)
    ]
    
    # Complete list of rows to write
    # Formatted as (is_header, data_tuple_or_header_string)
    all_rows = [
        # USER PORTAL SECTION
        (True, '  USER PORTAL'),
        (False, (1, 'USER PORTAL', 'Profile / Home', '/profile', 'User', 'Completed', 'Employee personal dashboard — shows profile info, department, role, recent activity, and quick links to key modules.', 'Done')),
        (False, (2, 'USER PORTAL', 'Notifications', '/notifications', 'User', 'Completed', 'Employee notification centre — inbox for system alerts, task assignments, leave decisions, meeting reminders, and broadcasts.', 'Done')),
        (False, (3, 'USER PORTAL', 'Notification Settings', '/notifications/settings', 'Sub-page', 'Completed', 'Allows employees to configure notification preferences — choose which event types trigger alerts.', 'Done')),
        (False, (4, 'USER PORTAL', 'Job Description Tool', '/tools/job-description', 'Sub-page', 'Completed', 'Tool for generating or previewing a formatted job description document for recruitment or role clarity purposes.', 'Done')),
        (False, (5, 'USER PORTAL', 'Projects', '/projects', 'Both', 'Not Started', 'Employee view of active and past projects — assigned projects, progress, team members, and deadlines. Shared with Admin.', 'TBD')),
        (False, (6, 'USER PORTAL', 'Tasks', '/tasks', 'Both', 'Completed', 'Personal task board — employees view, filter, and update tasks assigned to them. Shared with Admin.', 'Done')),
        (False, (7, 'USER PORTAL', 'Help Desk', '/help-desk', 'Both', 'Completed', 'IT & general support ticket system — employees raise tickets for issues or requests. Full status tracking included.', 'Done')),
        (False, (8, 'USER PORTAL', 'Reports Hub', '/reports', 'Both', 'Completed', 'Employee-facing reports portal — access to weekly meeting reports, KSS documents, minutes of meeting, and action trackers.', 'Done')),
        (False, (9, 'USER PORTAL', 'Weekly Reports', '/reports/weekly-reports', 'Sub-page', 'Completed', 'Lists auto-generated departmental weekly reports — employees view and download their department weekly performance summary.', 'Done')),
        (False, (10, 'USER PORTAL', 'Knowledge Sharing Session', '/reports/kss', 'Sub-page', 'Completed', 'Knowledge Sharing Session documents portal — view and download KSS presentations uploaded per meeting week.', 'Done')),
        (False, (11, 'USER PORTAL', 'Minutes of Meeting', '/reports/minutes-of-meeting', 'Sub-page', 'Completed', 'Access to uploaded Minutes of Meeting PDF documents — filterable by week and year.', 'Done')),
        (False, (12, 'USER PORTAL', 'Action Tracker', '/reports/action-tracker', 'Sub-page', 'Completed', 'Employee-facing action tracker — displays open, in-progress, and closed action items from general meetings.', 'Done')),
        (False, (13, 'USER PORTAL', 'Assets', '/assets', 'Both', 'Completed', 'Employee view of company assets — lists assets assigned to the individual or their department with status and acquisition details.', 'Done')),
        (False, (14, 'USER PORTAL', 'Payments', '/payments', 'User', 'Completed', 'Employee payments ledger — displays salary, allowances, and reimbursement records with status (due/paid/overdue).', 'Done')),
        (False, (15, 'USER PORTAL', 'Documentation', '/documentation', 'Both', 'Completed', 'Employee document library — browse company policies, SOPs, and internal documents. Shared with Admin.', 'Done')),
        (False, (16, 'USER PORTAL', 'Department Documents', '/documentation/department', 'Sub-page', 'Completed', 'Department-specific document repository — filtered view of documents belonging to the employee department.', 'Done')),
        (False, (17, 'USER PORTAL', 'Internal Documents', '/documentation/internal', 'Sub-page', 'Completed', 'Internal company documents — company-wide policies, HR circulars, compliance documents, and announcements.', 'Done')),
        (False, (18, 'USER PORTAL', 'Feedback', '/feedback', 'Both', 'Completed', 'Anonymous or named feedback submission — employees submit suggestions, concerns, or recognition. Shared with Admin.', 'Done')),
        (False, (19, 'USER PORTAL', 'Tools Hub', '/tools', 'Both', 'Completed', 'Self-service tool portal — signature generator, watermark tool, and reference letter generator. Shared with Admin.', 'Done')),
        (False, (20, 'USER PORTAL', 'Email Signature Generator', '/tools/signature', 'Sub-page', 'Completed', 'Generates a branded ACOB Lighting HTML email signature using the employee profile data — copyable for Outlook or Gmail.', 'Done')),
        (False, (21, 'USER PORTAL', 'Document Watermark Tool', '/tools/watermark', 'Sub-page', 'Completed', 'Applies a configurable watermark (confidential, draft, copy, etc.) to uploaded PDF documents with downloadable output.', 'Done')),
        (False, (22, 'USER PORTAL', 'Correspondence / Reference Letters', '/correspondence', 'User', 'Completed', 'Employee-facing correspondence portal — manage and generate official letters, letters of introduction, and other documents. Consolidated with Reference Letter Generator.', 'Done')),
        (False, (23, 'USER PORTAL', 'Leave', '/leave', 'User', 'Completed', 'Employee leave management — view leave balance, apply for leave, track approval status, and view leave history.', 'Done')),
        (False, (24, 'USER PORTAL', 'Leave Request', '/leave/request', 'Sub-page', 'Completed', 'Leave application form — employee selects leave type, date range, and provides justification. Triggers approval workflow.', 'Done')),
        (False, (25, 'USER PORTAL', 'Attendance', '/attendance', 'User', 'Completed', 'Employee attendance dashboard — view clock-in/clock-out records, total hours, and attendance status by date.', 'Done')),
        (False, (26, 'USER PORTAL', 'Attendance Records', '/attendance/records', 'Sub-page', 'Completed', 'Detailed historical attendance records — filterable by date range with export capability.', 'Done')),
        (False, (27, 'USER PORTAL', 'Fleet', '/fleet', 'User', 'Completed', 'Company fleet management for employees — redirects to general resources for vehicle and room booking.', 'Done')),
        (False, (28, 'USER PORTAL', 'Goals (PMS)', '/goals', 'User', 'Completed', 'Personal goal setting and tracking — employees set, update, and submit goals for approval with target values and due dates.', 'Done')),
        (False, (29, 'USER PORTAL', 'Reviews / Performance', '/reviews', 'User', 'Completed', 'Employee performance reviews — view submitted appraisals, ratings, strengths, improvement areas, and goals achieved.', 'Done')),
        (False, (30, 'USER PORTAL', 'Employee Directory', '/directory', 'User', 'Completed', 'Interactive directory showing all active employee profiles, departments, contact information, and roles.', 'Done')),
        (False, (31, 'USER PORTAL', 'Daily Activity Report', '/reports/daily-activity', 'User', 'Completed', 'Log and track daily work activity, submissions, and tasks performed.', 'Done')),
        (False, (32, 'USER PORTAL', 'Performance Hub (PMS)', '/pms', 'User', 'Completed', 'User-facing dashboard for goals, KPIs, reviews, and professional development.', 'Done')),
        (False, (33, 'USER PORTAL', 'PMS Goals', '/pms/goals', 'Sub-page', 'Completed', 'Setup, update, and track personal goals and objectives.', 'Done')),
        (False, (34, 'USER PORTAL', 'PMS KPIs', '/pms/kpi', 'Sub-page', 'Completed', 'View personal performance indicator targets and achievements.', 'Done')),
        (False, (35, 'USER PORTAL', 'PMS Peer Feedback', '/pms/peer-feedback', 'Sub-page', 'Completed', 'Submit and review peer performance feedback.', 'Done')),
        (False, (36, 'USER PORTAL', 'PMS Reviews', '/pms/reviews', 'Sub-page', 'Completed', 'Employee-facing performance evaluation reviews and self-appraisals.', 'Done')),
        (False, (37, 'USER PORTAL', 'PMS Attendance Link', '/pms/attendance', 'Sub-page', 'Completed', 'View attendance contribution to overall performance review.', 'Done')),
        (False, (38, 'USER PORTAL', 'PMS Behaviour / Core Values', '/pms/behaviour', 'Sub-page', 'Completed', 'Self-evaluation and review of core values alignment.', 'Done')),
        (False, (39, 'USER PORTAL', 'PMS CBT', '/pms/cbt', 'Sub-page', 'Completed', 'Access training and assessment tests for performance review.', 'Done')),
        (False, (40, 'USER PORTAL', 'PMS Development Plans', '/pms/development-plans', 'Sub-page', 'Completed', 'Track personal growth and professional development plan objectives.', 'Done')),
        (False, (41, 'USER PORTAL', 'General Meeting Reports Hub', '/reports/general-meeting', 'User', 'Completed', 'Main dashboard for general meeting outputs — weekly reports, KSS, minutes, and action items.', 'Done')),
        (False, (42, 'USER PORTAL', 'GM Action Tracker', '/reports/general-meeting/action-tracker', 'Sub-page', 'Completed', 'Detailed view of action items assigned during general meetings.', 'Done')),
        (False, (43, 'USER PORTAL', 'GM KSS', '/reports/general-meeting/kss', 'Sub-page', 'Completed', 'View and download Knowledge Sharing Session documents from general meetings.', 'Done')),
        (False, (44, 'USER PORTAL', 'GM Minutes', '/reports/general-meeting/minutes-of-meeting', 'Sub-page', 'Completed', 'Read and download official Minutes of Meeting PDFs from general meetings.', 'Done')),
        (False, (45, 'USER PORTAL', 'GM Weekly Reports', '/reports/general-meeting/weekly-reports', 'Sub-page', 'Completed', 'Access weekly reports from all departments submitted for general meetings.', 'Done')),
        (False, (46, 'USER PORTAL', 'Employee Resources', '/resources', 'User', 'Completed', 'General resources and quick links for employees, including company forms and files.', 'Done')),
        (False, (47, 'USER PORTAL', 'User Settings', '/settings', 'User', 'Completed', 'Account settings, password updates, and profile configuration.', 'Done')),
        (False, (48, 'USER PORTAL', 'Anniversary Signature Tool', '/tools/signature-anniversary', 'Sub-page', 'Completed', 'Generates custom anniversary-themed email signatures for employees.', 'Done')),

        # ADMIN PANEL SECTION
        (True, '  ADMIN PANEL'),
        (False, (49, 'ADMIN PANEL', 'Admin Dashboard', '/admin', 'Admin', 'Completed', 'Top-level admin overview — summary stats, quick links to all management modules, and system health indicators.', 'Done')),
        (False, (50, 'ADMIN PANEL', 'HR Management Hub', '/admin/hr', 'Admin', 'Completed', 'Central HR administration panel — overview of employees, pending leaves, departments, attendance, and performance KPIs.', 'Done')),
        (False, (51, 'ADMIN PANEL', 'HR – Employees List', '/admin/hr/employees', 'Sub-page', 'Completed', 'Full employee directory with filtering by department, status, and role — admin can view, edit, or deactivate profiles.', 'Done')),
        (False, (52, 'ADMIN PANEL', 'HR – Offboarding Conflicts', '/admin/hr/employees/offboarding-conflicts', 'Sub-page', 'Not Started', 'Flags employees with unresolved offboarding conflicts — e.g. pending leaves, open tickets, or incomplete handovers.', 'TBD')),
        (False, (53, 'ADMIN PANEL', 'HR – Departments', '/admin/hr/departments', 'Sub-page', 'Completed', 'Department management — create, edit, and manage company departments; view headcount and department hierarchy.', 'Done')),
        (False, (54, 'ADMIN PANEL', 'HR – Leave Management', '/admin/hr/leave', 'Sub-page', 'Completed', 'HR leave approval queue — review pending leave requests, approve or reject with comments, and manage calendar conflicts.', 'Done')),
        (False, (55, 'ADMIN PANEL', 'HR – Fleet Management', '/admin/hr/fleet', 'Sub-page', 'Completed', 'Admin fleet management — register vehicles, assign drivers, manage bookings, and track vehicle maintenance records.', 'Done')),
        (False, (56, 'ADMIN PANEL', 'HR – Office Location', '/admin/hr/office-location', 'Sub-page', 'Completed', 'Manage registered office locations — define geo-fence boundaries for attendance clock-in validation.', 'Done')),
        (False, (57, 'ADMIN PANEL', 'HR – Attendance Scoping', '/admin/hr/attendance', 'Sub-page', 'Completed', 'Org-wide attendance dashboard — real-time clock-in roster, exceptions, geo-fencing, and log details.', 'Done')),
        (False, (58, 'ADMIN PANEL', 'HR – Attendance Records', '/admin/hr/attendance/records', 'Sub-page', 'Completed', 'Full historical search and export of employee attendance logs.', 'Done')),
        (False, (59, 'ADMIN PANEL', 'HR – Attendance Remote Access', '/admin/hr/attendance/remote-access', 'Sub-page', 'Completed', 'Approve/revoke remote clock-in permissions for specific employees.', 'Done')),
        (False, (60, 'ADMIN PANEL', 'HR – Site Locations', '/admin/hr/site-locations', 'Sub-page', 'Completed', 'Define geofences and IP restrictions for site-based employee clock-in.', 'Done')),
        (False, (61, 'ADMIN PANEL', 'HR – Performance Review (PMS)', '/admin/hr/pms', 'Sub-page', 'Completed', 'Central panel for organizing performance review cycles, KPIs, core values, and appraisals.', 'Done')),
        (False, (62, 'ADMIN PANEL', 'HR – PMS Cycles', '/admin/hr/pms/cycles', 'Sub-page', 'Completed', 'Create, schedule, and open/close performance review cycles.', 'Done')),
        (False, (63, 'ADMIN PANEL', 'HR – PMS Goals', '/admin/hr/pms/goals', 'Sub-page', 'Completed', 'Review and approve employee-submitted goals and targets.', 'Done')),
        (False, (64, 'ADMIN PANEL', 'HR – PMS KPIs', '/admin/hr/pms/kpi', 'Sub-page', 'Completed', 'Define global and department-specific KPIs and weights.', 'Done')),
        (False, (65, 'ADMIN PANEL', 'HR – PMS Reviews Appraisals', '/admin/hr/pms/reviews', 'Sub-page', 'Completed', 'Access all employee appraisals, manage review assignments, and sign off reviews.', 'Done')),
        (False, (66, 'ADMIN PANEL', 'HR – PMS Competencies', '/admin/hr/pms/competencies', 'Sub-page', 'Completed', 'Manage professional competency standards per department and role.', 'Done')),
        (False, (67, 'ADMIN PANEL', 'HR – PMS Behaviour / Values', '/admin/hr/pms/behaviour', 'Sub-page', 'Completed', 'Setup appraisal criteria for core values and soft skills.', 'Done')),
        (False, (68, 'ADMIN PANEL', 'HR – PMS Attendance Score', '/admin/hr/pms/attendance', 'Sub-page', 'Completed', 'Define attendance score contribution and rules for performance evaluations.', 'Done')),
        (False, (69, 'ADMIN PANEL', 'HR – PMS CBT Management', '/admin/hr/pms/cbt', 'Sub-page', 'Completed', 'Manage computer-based tests, question banks, and review candidate scores.', 'Done')),
        (False, (70, 'ADMIN PANEL', 'HR – PMS Analytics Dashboard', '/admin/hr/pms/analytics', 'Sub-page', 'Completed', 'Overall performance rating distribution, department scores, and growth analytics.', 'Done')),
        (False, (71, 'ADMIN PANEL', 'HR – PMS Development Plans', '/admin/hr/pms/development-plans', 'Sub-page', 'Completed', 'Monitor employee professional development plans and training compliance.', 'Done')),
        (False, (72, 'ADMIN PANEL', 'HR – Daily Activity Logs', '/admin/hr/reports/daily-activity', 'Sub-page', 'Completed', 'Review daily work logs submitted by employees across the company.', 'Done')),
        (False, (73, 'ADMIN PANEL', 'HR – Resources Manager', '/admin/hr/resources', 'Sub-page', 'Completed', 'Upload company policies, onboarding guides, and templates for staff.', 'Done')),
        (False, (74, 'ADMIN PANEL', 'Finance Hub', '/admin/finance', 'Admin', 'Completed', 'Finance overview dashboard — summary of invoices, bills, payments, and financial health indicators across all departments.', 'Done')),
        (False, (75, 'ADMIN PANEL', 'Finance – Invoices', '/admin/finance/invoices', 'Sub-page', 'Completed', 'Invoice management — create, view, and track outgoing invoices; status tracking from draft to paid.', 'Done')),
        (False, (76, 'ADMIN PANEL', 'Finance – Bills', '/admin/finance/bills', 'Sub-page', 'Completed', 'Vendor bills management — log, track, and approve bills; status tracking from draft to paid.', 'Done')),
        (False, (77, 'ADMIN PANEL', 'Finance – Payments Ledger', '/admin/finance/payments', 'Sub-page', 'Completed', 'Admin payments ledger — manage all employee or vendor payment records; filter by status, department, or date.', 'Done')),
        (False, (78, 'ADMIN PANEL', 'Finance – Payments by Dept', '/admin/finance/payments/departments', 'Sub-page', 'Completed', 'Department-level payment breakdown — view total payment obligations and disbursements grouped by department.', 'Done')),
        (False, (79, 'ADMIN PANEL', 'Finance – Reports', '/admin/finance/reports', 'Sub-page', 'Completed', 'Financial reporting module — generates summary and detailed financial reports exportable to PDF or spreadsheet.', 'Done')),
        (False, (80, 'ADMIN PANEL', 'Projects (Admin)', '/admin/projects', 'Both', 'Not Started', 'Admin-side project management — overview of all company projects, team assignment, and progress monitoring.', 'TBD')),
        (False, (81, 'ADMIN PANEL', 'Tasks (Admin)', '/admin/tasks', 'Both', 'Completed', 'Admin task management — view and manage all tasks across all employees and departments; reassign and escalate.', 'Done')),
        (False, (82, 'ADMIN PANEL', 'Help Desk (Admin)', '/admin/help-desk', 'Both', 'Completed', 'Admin support ticket management — full queue of all raised tickets, assignment to staff, status updates, and SLA tracking.', 'Done')),
        (False, (83, 'ADMIN PANEL', 'Reports Admin Hub', '/admin/reports', 'Both', 'Completed', 'Admin reports management centre — upload, manage, and broadcast meeting reports, KSS, minutes, and action points.', 'Done')),
        (False, (84, 'ADMIN PANEL', 'Reports – Weekly Reports', '/admin/reports/weekly-reports', 'Sub-page', 'Completed', 'Admin-side weekly report management — view all departmental weekly reports, filter by week/year, and export.', 'Done')),
        (False, (85, 'ADMIN PANEL', 'Reports – KSS Upload', '/admin/reports/kss', 'Sub-page', 'Completed', 'Upload and manage Knowledge Sharing Session files — version control, presenter tagging, and per-week tracking.', 'Done')),
        (False, (86, 'ADMIN PANEL', 'Reports – Minutes Upload', '/admin/reports/minutes-of-meeting', 'Sub-page', 'Completed', 'Upload and manage Minutes of Meeting PDF documents — versioned per meeting week with current-version tracking.', 'Done')),
        (False, (87, 'ADMIN PANEL', 'Reports – Action Tracker', '/admin/reports/action-tracker', 'Sub-page', 'Completed', 'Admin action tracker management — create, update, and close action items; assign responsible parties and deadlines.', 'Done')),
        (False, (88, 'ADMIN PANEL', 'Reports – Meeting Documents', '/admin/reports/meeting-documents', 'Sub-page', 'Completed', 'Master view of all meeting documents per week — KSS, minutes, and action point files with download and version info.', 'Done')),
        (False, (89, 'ADMIN PANEL', 'Reports – Weekly Summary Mail', '/admin/reports/mail', 'Sub-page', 'Completed', 'Weekly meeting summary broadcast — select recipients, content types (report/KSS/minutes/action tracker), send or schedule emails with PDF attachments.', 'Done')),
        (False, (90, 'ADMIN PANEL', 'Correspondence Administration', '/admin/correspondence', 'Admin', 'Completed', 'Admin correspondence dashboard — generate and approve reference/intro letters, manage templates, and upload documents. Consolidated with Tools - Reference Generator.', 'Done')),
        (False, (91, 'ADMIN PANEL', 'Notifications (Admin)', '/admin/notifications', 'Both', 'Completed', 'Admin notification management centre — send broadcasts, manage meeting notifications, and configure reminder rules.', 'Done')),
        (False, (92, 'ADMIN PANEL', 'Communications – Broadcast', '/admin/communications/broadcast', 'Sub-page', 'Completed', 'Email/notification broadcast centre — compose and send company-wide or targeted messages to employees. Consolidated from Notifications - Broadcast.', 'Done')),
        (False, (93, 'ADMIN PANEL', 'Communications – Meetings', '/admin/communications/meetings', 'Sub-page', 'Completed', 'Meeting communications management — view meeting schedule, manage invites, and coordinate meeting correspondence.', 'Done')),
        (False, (94, 'ADMIN PANEL', 'Communications – Meeting Mail', '/admin/communications/meetings/mail', 'Sub-page', 'Completed', 'Send meeting-specific emails — compose and dispatch formal meeting invitations and follow-ups.', 'Done')),
        (False, (95, 'ADMIN PANEL', 'Communications – Reminders', '/admin/communications/meetings/reminders', 'Sub-page', 'Completed', 'Schedule automated meeting reminder emails — configure timing and recipient list for recurring reminders.', 'Done')),
        (False, (96, 'ADMIN PANEL', 'Assets (Admin)', '/admin/assets', 'Both', 'Completed', 'Admin asset management — full company asset register with assignment, status tracking, and maintenance logs.', 'Done')),
        (False, (97, 'ADMIN PANEL', 'Assets – Issue Tracking', '/admin/assets/issues', 'Sub-page', 'Completed', 'Track reported asset issues and faults — log damage reports, assign repair tickets, and manage resolution status.', 'Done')),
        (False, (98, 'ADMIN PANEL', 'Documentation (Admin)', '/admin/documentation', 'Both', 'Completed', 'Admin document management — upload, categorise, and publish company documents visible to employees.', 'Done')),
        (False, (99, 'ADMIN PANEL', 'Documentation – Dept Docs', '/admin/documentation/department', 'Sub-page', 'Completed', 'Admin management of department-specific documents — upload, version, and control visibility per department.', 'Done')),
        (False, (100, 'ADMIN PANEL', 'Documentation – Internal Docs', '/admin/documentation/internal', 'Sub-page', 'Completed', 'Admin management of internal company-wide documents — HR policies, compliance docs, company announcements.', 'Done')),
        (False, (101, 'ADMIN PANEL', 'Feedback (Admin)', '/admin/feedback', 'Both', 'Completed', 'Admin feedback inbox — review employee feedback submissions; categorise, respond, and track resolution.', 'Done')),
        (False, (102, 'ADMIN PANEL', 'Audit Logs', '/admin/audit-logs', 'Admin', 'Completed', 'System-wide audit trail — logs all admin actions with actor, timestamp, entity type, and metadata.', 'Done')),
        (False, (103, 'ADMIN PANEL', 'Settings', '/admin/settings', 'Admin', 'Completed', 'Top-level admin settings hub — manage company config, mail settings, user roles, accounts, and system maintenance.', 'Done')),
        (False, (104, 'ADMIN PANEL', 'Settings – Company', '/admin/settings/company', 'Sub-page', 'Completed', 'Company profile settings — update company name, address, contact info, and branding used across the system.', 'Done')),
        (False, (105, 'ADMIN PANEL', 'Settings – Mail', '/admin/settings/mail', 'Sub-page', 'Completed', 'Email system configuration — manage mail provider settings, enable/disable automated emails per module, and test delivery.', 'Done')),
        (False, (106, 'ADMIN PANEL', 'Settings – Maintenance', '/admin/settings/maintenance', 'Sub-page', 'Completed', 'System maintenance mode toggle and scheduling — take the system offline with a custom user-facing message.', 'Done')),
        (False, (107, 'ADMIN PANEL', 'Settings – Roles & Permissions', '/admin/settings/roles', 'Sub-page', 'Completed', 'Role-based access control (RBAC) management — define roles, assign module permissions, and configure access scopes.', 'Done')),
        (False, (108, 'ADMIN PANEL', 'Settings – User Management', '/admin/settings/users', 'Sub-page', 'Completed', 'Manage all system user accounts — view, activate/deactivate, reset passwords, and change roles. Consolidated from Settings - Invite User.', 'Done')),
        (False, (109, 'ADMIN PANEL', 'Inventory Hub', '/admin/inventory', 'Admin', 'Completed', 'Admin inventory management — overview of all products, stock levels, warehouse locations, and movement history.', 'Done')),
        (False, (110, 'ADMIN PANEL', 'Inventory – Products', '/admin/inventory/products', 'Sub-page', 'Completed', 'Full product catalogue — list, filter, and manage all inventory items with SKU, category, and stock quantity.', 'Done')),
        (False, (111, 'ADMIN PANEL', 'Inventory – Categories', '/admin/inventory/categories', 'Sub-page', 'Completed', 'Product category management — create and organise inventory categories and sub-categories.', 'Done')),
        (False, (112, 'ADMIN PANEL', 'Inventory – Warehouses', '/admin/inventory/warehouses', 'Sub-page', 'Completed', 'Warehouse/storage location management — register warehouses, zones, and bin locations for stock tracking.', 'Done')),
        (False, (113, 'ADMIN PANEL', 'Inventory – Stock Movements', '/admin/inventory/movements', 'Sub-page', 'Completed', 'Stock movement log — tracks all inbound and outbound inventory movements with date, quantity, and reference.', 'Done')),
        (False, (114, 'ADMIN PANEL', 'Purchasing Hub', '/admin/purchasing', 'Admin', 'Completed', 'Procurement management dashboard — overview of active purchase orders, supplier performance, and pending receipts.', 'Done')),
        (False, (115, 'ADMIN PANEL', 'Purchasing – Orders', '/admin/purchasing/orders', 'Sub-page', 'Completed', 'Purchase order management — list, filter, and track all purchase orders from creation through to receipt.', 'Done')),
        (False, (116, 'ADMIN PANEL', 'Purchasing – Suppliers', '/admin/purchasing/suppliers', 'Sub-page', 'Completed', 'Supplier directory — register and manage vendors with contact details, payment terms, and performance ratings.', 'Done')),
        (False, (117, 'ADMIN PANEL', 'Purchasing – Receipts', '/admin/purchasing/receipts', 'Sub-page', 'Completed', 'Goods receipt management — log and confirm delivery of purchased items against purchase orders.', 'Done')),
        (False, (118, 'ADMIN PANEL', 'Admin Job Descriptions', '/admin/job-descriptions', 'Admin', 'Completed', 'Admin management of company job descriptions — create, edit, assign, and publish JDs for all roles.', 'Done')),
        (False, (119, 'ADMIN PANEL', 'GM Reports Hub Admin', '/admin/reports/general-meeting', 'Sub-page', 'Completed', 'Admin-side general meeting reports dashboard — weekly reports, KSS, minutes, and action items.', 'Done')),
        (False, (120, 'ADMIN PANEL', 'GM Action Tracker Admin', '/admin/reports/general-meeting/action-tracker', 'Sub-page', 'Completed', 'Admin action items manager for general meetings — create, assign, and track meeting action points.', 'Done')),
        (False, (121, 'ADMIN PANEL', 'GM KSS Admin', '/admin/reports/general-meeting/kss', 'Sub-page', 'Completed', 'Upload and catalog Knowledge Sharing Session presentations for general meetings.', 'Done')),
        (False, (122, 'ADMIN PANEL', 'GM Minutes Admin', '/admin/reports/general-meeting/minutes-of-meeting', 'Sub-page', 'Completed', 'Upload and publish official general meeting minutes.', 'Done')),
        (False, (123, 'ADMIN PANEL', 'GM Weekly Reports Admin', '/admin/reports/general-meeting/weekly-reports', 'Sub-page', 'Completed', 'Review and merge weekly reports from all departments.', 'Done')),

        # DEVELOPER TOOLS SECTION
        (True, '  DEVELOPER TOOLS'),
        (False, (124, 'DEVELOPER TOOLS', 'Dev Dashboard', '/admin/dev', 'Admin (Dev)', 'Completed', 'Developer-only section overview — quick links to all diagnostic and monitoring tools. Consolidated from Dev Maintenance.', 'Done')),
        (False, (125, 'DEVELOPER TOOLS', 'Login Logs', '/admin/dev/login-logs', 'Admin (Dev)', 'Completed', 'System login activity log — all user sign-in events with timestamp, IP, device, and success/failure status.', 'Done')),
        (False, (126, 'DEVELOPER TOOLS', 'Role Escalations', '/admin/dev/role-escalations', 'Admin (Dev)', 'Completed', 'Tracks temporary role escalation events — when users were granted elevated permissions and for how long.', 'Done')),
        (False, (127, 'DEVELOPER TOOLS', 'Security Events', '/admin/dev/security-events', 'Admin (Dev)', 'Completed', 'Security incident log — records suspicious actions, failed authentication attempts, and policy violations.', 'Done')),
        (False, (128, 'DEVELOPER TOOLS', 'Tests', '/admin/dev/tests', 'Admin (Dev)', 'Completed', 'Built-in integration test runner — execute end-to-end system tests and view test results and diagnostics.', 'Done')),
        (False, (129, 'DEVELOPER TOOLS', 'UI Error Monitor', '/admin/dev/ui-errors', 'Admin (Dev)', 'Completed', 'Frontend error tracking — logs client-side JavaScript errors and React boundary exceptions with stack traces.', 'Done')),
        (False, (130, 'DEVELOPER TOOLS', 'Impersonation Tool', '/admin/dev/impersonation', 'Admin (Dev)', 'Completed', 'Developer diagnostic tool for impersonating employee accounts to debug specific user-level issues.', 'Done')),
        (False, (131, 'DEVELOPER TOOLS', 'Acobot Diagnostics', '/admin/dev/acobot', 'Admin (Dev)', 'Completed', 'AI system diagnostics — monitor chatbot context, prompt states, and test response parameters.', 'Done')),

        # AUTH PAGES SECTION
        (True, '  AUTH PAGES'),
        (False, (132, 'AUTH PAGES', 'Login', '/auth/login', 'Auth', 'Completed', 'Employee login page — email and password authentication with session management via Supabase Auth.', 'Done')),
        (False, (133, 'AUTH PAGES', 'Sign Up', '/auth/sign-up', 'Auth', 'Completed', 'New account registration — used during employee onboarding via invite link.', 'Done')),
        (False, (134, 'AUTH PAGES', 'Sign Up Success', '/auth/sign-up-success', 'Auth', 'Completed', 'Post-registration confirmation page — confirms successful account creation and guides employee to next steps.', 'Done')),
        (False, (135, 'AUTH PAGES', 'Forgot Password', '/auth/forgot-password', 'Auth', 'Completed', 'Password reset request — employee submits their email to receive a secure password reset link.', 'Done')),
        (False, (136, 'AUTH PAGES', 'Reset Password', '/auth/reset-password', 'Auth', 'Completed', 'Password reset form — employee sets a new password using the secure token received via email.', 'Done')),
        (False, (137, 'AUTH PAGES', 'Set Password', '/auth/set-password', 'Auth', 'Completed', 'Initial password setup page — used by newly invited employees to set their first password.', 'Done')),
        (False, (138, 'AUTH PAGES', 'Setup Account / Onboarding', '/auth/setup-account', 'Auth', 'Completed', 'New employee account setup wizard — collects profile details, department, and preferences during first-time login.', 'Done')),
        (False, (139, 'AUTH PAGES', 'Auth Error Page', '/auth/error', 'Auth', 'Completed', 'Auth error boundary page — displayed when authentication fails or tokens are invalid/expired.', 'Done')),

        # DEPARTMENT CONSOLE SECTION
        (True, '  DEPARTMENT CONSOLE'),
        (False, (140, 'DEPARTMENT CONSOLE', 'Department Dashboard', '/dept/[dept_id]', 'Department', 'Completed', 'Department lead console — summary statistics, departmental headcount, quick actions, and recent activities.', 'Done')),
        (False, (141, 'DEPARTMENT CONSOLE', 'Department HR Hub', '/dept/[dept_id]/hr', 'Department', 'Completed', 'Manager view of department employee profiles, roles, and status.', 'Done')),
        (False, (142, 'DEPARTMENT CONSOLE', 'Dept HR – Attendance Daily', '/dept/[dept_id]/hr/attendance', 'Sub-page', 'Completed', 'Manager view of department daily attendance, clock-in records, and roster.', 'Done')),
        (False, (143, 'DEPARTMENT CONSOLE', 'Dept HR – Attendance Logs', '/dept/[dept_id]/hr/attendance/records', 'Sub-page', 'Completed', 'Filtered view of department daily attendance history and export capability.', 'Done')),
        (False, (144, 'DEPARTMENT CONSOLE', 'Dept HR – Leave Approvals', '/dept/[dept_id]/hr/leave', 'Sub-page', 'Completed', 'Supervisor review for department leave requests — reliever coverage and schedule conflict checking.', 'Done')),
        (False, (145, 'DEPARTMENT CONSOLE', 'Dept HR – Performance (PMS)', '/dept/[dept_id]/hr/pms', 'Sub-page', 'Completed', 'Departmental goals, appraisal review queue, and performance stats.', 'Done')),
        (False, (146, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Analytics', '/dept/[dept_id]/hr/pms/analytics', 'Sub-page', 'Completed', 'Departmental performance review analytics, rating distributions, and targets.', 'Done')),
        (False, (147, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Goals Queue', '/dept/[dept_id]/hr/pms/goals', 'Sub-page', 'Completed', 'Review and approve department employee goals and target configurations.', 'Done')),
        (False, (148, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Reviews Queue', '/dept/[dept_id]/hr/pms/reviews', 'Sub-page', 'Completed', 'Supervisor performance review evaluation queue for department employees.', 'Done')),
        (False, (149, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Competencies', '/dept/[dept_id]/hr/pms/competencies', 'Sub-page', 'Completed', 'View and manage professional competency standards for department roles.', 'Done')),
        (False, (150, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Behaviour', '/dept/[dept_id]/hr/pms/behaviour', 'Sub-page', 'Completed', 'Assess department employees on core values and soft skills compliance.', 'Done')),
        (False, (151, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Attendance Score', '/dept/[dept_id]/hr/pms/attendance', 'Sub-page', 'Completed', 'Monitor department employees daily attendance contribution to performance score.', 'Done')),
        (False, (152, 'DEPARTMENT CONSOLE', 'Dept HR – PMS CBT Results', '/dept/[dept_id]/hr/pms/cbt', 'Sub-page', 'Completed', 'View department candidate scores for training and compliance tests.', 'Done')),
        (False, (153, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Dev Plans', '/dept/[dept_id]/hr/pms/development-plans', 'Sub-page', 'Completed', 'Track department employee growth and training plan compliance.', 'Done')),
        (False, (154, 'DEPARTMENT CONSOLE', 'Dept HR – Daily Activity Logs', '/dept/[dept_id]/hr/reports/daily-activity', 'Sub-page', 'Completed', 'Review daily work logs submitted by employees within the department.', 'Done')),
        (False, (155, 'DEPARTMENT CONSOLE', 'Department Tasks', '/dept/[dept_id]/tasks', 'Department', 'Completed', 'Team task board — create, assign, and track tasks for the department.', 'Done')),
        (False, (156, 'DEPARTMENT CONSOLE', 'Department Finance Hub', '/dept/[dept_id]/finance', 'Department', 'Completed', 'Track department-specific invoices, bills, budgets, and payments.', 'Done')),
        (False, (157, 'DEPARTMENT CONSOLE', 'Dept Finance – Bills', '/dept/[dept_id]/finance/bills', 'Sub-page', 'Completed', 'Review, approve, and track vendor bills logged for the department.', 'Done')),
        (False, (158, 'DEPARTMENT CONSOLE', 'Dept Finance – Invoices', '/dept/[dept_id]/finance/invoices', 'Sub-page', 'Completed', 'Manage outgoing invoices generated by the department.', 'Done')),
        (False, (159, 'DEPARTMENT CONSOLE', 'Dept Finance – Payments', '/dept/[dept_id]/finance/payments', 'Sub-page', 'Completed', 'Department-level payment obligations and disbursement records.', 'Done')),
        (False, (160, 'DEPARTMENT CONSOLE', 'Dept Finance – Reports', '/dept/[dept_id]/finance/reports', 'Sub-page', 'Completed', 'Financial reporting module — generates departmental summaries and detailed expense sheets.', 'Done')),
        (False, (161, 'DEPARTMENT CONSOLE', 'Department Document Library', '/dept/[dept_id]/documentation', 'Department', 'Completed', 'Manage files, policies, and standard operating procedures specific to the department.', 'Done')),
        (False, (162, 'DEPARTMENT CONSOLE', 'Dept Docs – Internal Docs', '/dept/[dept_id]/documentation/internal', 'Sub-page', 'Completed', 'Read-only access to company-wide internal circulars and policies within console.', 'Done')),
        (False, (163, 'DEPARTMENT CONSOLE', 'Dept Docs – Departmental', '/dept/[dept_id]/documentation/department', 'Sub-page', 'Completed', 'Upload, manage, and download documents and SOPs specific to the department.', 'Done')),
        (False, (164, 'DEPARTMENT CONSOLE', 'Department Help Desk', '/dept/[dept_id]/help-desk', 'Department', 'Completed', 'Handle support tickets assigned to or raised by the department.', 'Done')),
        (False, (165, 'DEPARTMENT CONSOLE', 'Department Reports Hub', '/dept/[dept_id]/reports', 'Department', 'Completed', 'Manage department weekly meeting reports, minutes, and action trackers.', 'Done')),
        (False, (166, 'DEPARTMENT CONSOLE', 'Dept Reports – Weekly Reports', '/dept/[dept_id]/reports/weekly', 'Sub-page', 'Completed', 'Submit and edit department weekly reports for general meetings.', 'Done')),
        (False, (167, 'DEPARTMENT CONSOLE', 'Department Feedback', '/dept/[dept_id]/feedback', 'Department', 'Completed', 'Review feedback from team members within the department.', 'Done')),
        (False, (168, 'DEPARTMENT CONSOLE', 'Department Communications', '/dept/[dept_id]/communications', 'Department', 'Completed', 'Send internal communications and announcements to department members.', 'Done')),
        (False, (169, 'DEPARTMENT CONSOLE', 'Department Assets', '/dept/[dept_id]/assets', 'Department', 'Completed', 'View and request company assets assigned to the department.', 'Done'))
    ]

    current_row = 5
    print(f"Writing {len(all_rows)} rows of content...")
    
    for item in all_rows:
        is_header, data = item
        if is_header:
            # Recreate section header row
            ws.row_dimensions[current_row].height = 22.05
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            
            # Write value to A and style all merged cells to render correctly
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_header
                cell.border = thin_border
                if col == 1:
                    cell.value = data
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            # Recreate data row
            ws.row_dimensions[current_row].height = 34.05
            
            # Extract data tuple
            idx, section, title, url_path, location, status, desc, timeline = data
            
            # Cells values
            values = [idx, section, title, url_path, location, status, desc, timeline]
            
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.value = values[col - 1]
                cell.border = thin_border
                cell.alignment = alignments[col - 1]
                
                # Default background fill
                cell.fill = fill_white
                
                # Fonts and custom fills
                if col == 3: # Title (Column C)
                    cell.font = font_title
                elif col == 6: # Status (Column F)
                    # Status styling
                    s_val = values[col - 1]
                    if s_val in status_styles:
                        cell.font = status_styles[s_val]['font']
                        cell.fill = status_styles[s_val]['fill']
                    else:
                        cell.font = font_regular
                else:
                    cell.font = font_regular
                    
        current_row += 1

    print(f"Saving workspace copy to {WORKSPACE_XLSX}...")
    wb.save(WORKSPACE_XLSX)
    print("Workspace copy saved successfully!")
    
    # 3. Copy to Downloads folder
    print(f"Copying workspace file to {DOWNLOADS_XLSX}...")
    try:
        shutil.copy(WORKSPACE_XLSX, DOWNLOADS_XLSX)
        print(f"Successfully updated and copied tracker to {DOWNLOADS_XLSX}!")
    except Exception as e:
        print(f"Error copying to downloads: {e}")

if __name__ == '__main__':
    main()
