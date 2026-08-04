import openpyxl
import datetime
import dotenv
import psycopg2
import re

env = dotenv.dotenv_values('.env.local')
db_url = env.get('SUPABASE_POSTGRES_URL_NON_POOLING') or env.get('SUPABASE_POSTGRES_URL')
conn = psycopg2.connect(db_url)
cur = conn.cursor()

cur.execute('SELECT id, full_name, first_name, last_name FROM profiles;')
db_profiles = cur.fetchall()
prof_by_id = {p[0]: p[1] for p in db_profiles}

# Load Excel
wb = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\ACOB PETTY CASH BOOKS.xlsx', data_only=True)
ws_lakita = wb['LAKITA FOOD 2026']
ws_ded = wb['FOOD DEDUCTION 2026']

headers_l = [ws_lakita.cell(1, c).value for c in range(1, ws_lakita.max_column + 1)]
headers_d = [ws_ded.cell(1, c).value for c in range(1, ws_ded.max_column + 1)]

MANUAL_NAME_MAP = {
    'shirley': 'Shirley Jackreece',
    'john': 'John Dangana',
    'susan eze': 'Susan Eze',
    'emmanuel': 'Emmanuel Ibanga',
    'vanessa': 'Vanessa Lawrence-Ukaegbu',
    'thomas': 'Thomas Olobo',
    'philip': 'Philip Yakubu',
    'busayo': 'Busayo Kadiri',
    'edward': 'Edward Atoshi',
    'andrew': 'Andrew Inegbedion',
    'kennedy': 'Kennedy Odenigbo',
    'tansi': 'Tansi',
    'dennis': 'Dennis Innug',
    'lawrence': 'Lawrence Adukwu',
    'favour': 'Favour Onuoha-Eke',
    'jessica': 'Jessica Egeonu',
    'bawa': 'Alhamdu Bawa',
    'samila': 'Agya Samila',
    'elijah': 'Elijah Isah',
    'tunde': 'Tunde Ajayi',
    'patrick': 'Patrick Prosper',
    'arc emmanuel': 'Emmanuel Chinedu',
    'tochucku': 'Tochukwu Nnadozie',
    'samad': 'Abdulsamad Danmusa',
    'abiodun': 'Sefui Abiodun',
    'chibuikim': 'Chibuikem Ilonze',
    'victor': 'Victor Onyeka',
    'usman': 'Usman Haruna',
    'habeeb': 'Habeeb',
    'habib': 'Habeeb',
    'suraj': 'Surajo Idris',
    'caleb': 'Caleb Obiechina',
    'daniel': 'Daniel Osa-Egonwa',
    'anointing': 'Anointed Emoghene',
    'oghenerune': 'Oghenerune Orhorhomuke',
    'rune': 'Oghenerune Orhorhomuke',
    'benson': 'Benson Ademoye',
    'sommy': 'Sommy',
    'ayoola': 'Peter Ayoola',
    'gregory': 'Gregory',
    'onyeka': 'Onyeka Eze',
    'oyneka': 'Onyeka Eze',
    'khleo': 'Kleopatra Aiyede',
    'cleopatra': 'Kleopatra Aiyede',
    'abdul': 'Abdulmalik Abdulkarim',
    'haron': 'Ode Haron',
    'peace': 'Peace Dogara',
    'new daniel': 'Istifanus Daniel',
    'seun': 'Oluwaseun Awotona',
    'tobi': 'Oluwatobi Oladele',
    'noble': 'Noble'
}

def normalize(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s)).strip().lower()

def find_profile_for_header(header):
    norm = normalize(header)
    if not norm or norm == 'total' or norm == 'names' or norm == 'date': 
        return None
    if norm in MANUAL_NAME_MAP:
        mapped_name = MANUAL_NAME_MAP[norm]
        for p in db_profiles:
            if normalize(p[1]) == normalize(mapped_name):
                return p
    for p in db_profiles:
        if normalize(p[1]) == norm or normalize(p[2]) == norm or normalize(p[3]) == norm:
            return p
    return None

emp_map = {}
for c in range(3, 51):
    h = headers_l[c-1] or headers_d[c-1]
    p = find_profile_for_header(h)
    if p:
        emp_map[c] = p

def parse_date_fixed(val):
    dt = None
    if isinstance(val, datetime.datetime):
        dt = val.date()
    elif isinstance(val, datetime.date):
        dt = val
    elif isinstance(val, str):
        val = val.strip()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
            try:
                dt = datetime.datetime.strptime(val, fmt).date()
                break
            except ValueError:
                pass
    if not dt:
        return None
    if dt.year == 2025:
        dt = datetime.date(2026, dt.month, dt.day)
    if dt == datetime.date(2026, 1, 31):
        dt = datetime.date(2026, 1, 30)
    return dt

jan_dates_l = []
for r in range(1, 20):
    c2 = ws_lakita.cell(r, 2).value
    dt = parse_date_fixed(c2)
    if dt and dt.month == 1 and dt.year == 2026:
        raw_date_str = str(c2)
        raw_dow = str(ws_lakita.cell(r, 1).value).strip()
        jan_dates_l.append((r, dt, raw_dow, raw_date_str))

jan_excel = {}
for r, dt, dow, raw in jan_dates_l:
    for c in range(3, 51):
        vl = ws_lakita.cell(r, c).value or 0
        vd = ws_ded.cell(r, c).value or 0
        if (isinstance(vl, (int, float)) and vl > 0) or (isinstance(vd, (int, float)) and vd > 0):
            jan_excel[(dt, c)] = (vl, vd)

cur.execute("SELECT date, user_id FROM attendance_lunch_log WHERE date >= '2026-01-01' AND date <= '2026-01-31';")
db_jan_logs = set(cur.fetchall())

print(f"Jan Excel Total Entries: {len(jan_excel)}")
print(f"Jan DB Total Logs: {len(db_jan_logs)}")

# Category 1: Date & Year Typos
print("\n--- 1. DATE & YEAR TYPOS IN EXCEL (JANUARY 2026) ---")
for r, dt, dow, raw in jan_dates_l:
    if '2025' in raw:
        print(f"Row {r:2d} | Label: {dow:<9} | Typed in Excel: '{raw}' | Corrected Date: {dt}")

# Category 2: Cross-sheet Mismatches (Lakita vs Deduction)
print("\n--- 2. LAKITA FOOD VS FOOD DEDUCTION MISMATCHES ---")
mismatch_count = 0
for (dt, c), (vl, vd) in sorted(jan_excel.items()):
    if vl != vd:
        mismatch_count += 1
        emp_name = headers_l[c-1]
        print(f"Date: {dt} | Col {c} ({emp_name}): Lakita Vendor=NGN {vl:,.0f} vs Payroll Deduction=NGN {vd:,.0f}")
print(f"Total Sheet Mismatches in Jan: {mismatch_count}")

# Category 3: Recorded in Excel but Missing in DB
print("\n--- 3. RECORDED IN EXCEL BUT MISSING IN DATABASE ---")
missing_in_db = []
for (dt, c), (vl, vd) in sorted(jan_excel.items()):
    if c in emp_map:
        uid = emp_map[c][0]
        full_name = emp_map[c][1]
        if (dt, uid) not in db_jan_logs:
            missing_in_db.append((dt, full_name, vl, vd))
            print(f"Date: {dt} | Employee: {full_name:<25} | Excel Lakita: NGN {vl:,.0f} | Excel Deduction: NGN {vd:,.0f}")
print(f"Total Missing in DB for Jan: {len(missing_in_db)}")

# Category 4: Recorded in DB but Missing in Excel
print("\n--- 4. RECORDED IN DATABASE BUT MISSING IN EXCEL ---")
excel_uids_by_date = set((dt, emp_map[c][0]) for (dt, c) in jan_excel.keys() if c in emp_map)
extra_in_db = []
for (dt, uid) in sorted(db_jan_logs):
    if (dt, uid) not in excel_uids_by_date:
        name = prof_by_id.get(uid, uid)
        extra_in_db.append((dt, name))
        print(f"Date: {dt} | Employee: {name}")
print(f"Total Extra in DB for Jan: {len(extra_in_db)}")

conn.close()
