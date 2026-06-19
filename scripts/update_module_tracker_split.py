import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define paths
WORKSPACE_XLSX = 'C:/Users/IT_COMMS/GitHubProjects/ACOB-Signature-Creator/ACOB_ERP_Module_Tracker.xlsx'
DOWNLOADS_XLSX = 'C:/Users/IT_COMMS/Downloads/ACOB_ERP_Module_Tracker.xlsx'

# Define standard styling helpers
font_regular = Font(name='Arial', size=9, color='FF374151')
font_title = Font(name='Arial', size=9, bold=True, color='FF111827')
font_header = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
font_legend = Font(name='Arial', size=9, bold=True)

fill_white = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
fill_header = PatternFill(fill_type='solid', fgColor='FF1A3D28')

status_styles = {
    'Completed': {
        'font': Font(name='Arial', size=9, bold=True, color='FF166534'),
        'fill': PatternFill(fill_type='solid', fgColor='FFDCFCE7')
    },
    'In Progress': {
        'font': Font(name='Arial', size=9, bold=True, color='FF854D0E'),
        'fill': PatternFill(fill_type='solid', fgColor='FFFEF9C3')
    },
    'Not Started': {
        'font': Font(name='Arial', size=9, bold=True, color='FF991B1B'),
        'fill': PatternFill(fill_type='solid', fgColor='FFFEE2E2')
    }
}

thin_border = Border(
    left=Side(style='thin', color='FFD1D5DB'),
    right=Side(style='thin', color='FFD1D5DB'),
    top=Side(style='thin', color='FFD1D5DB'),
    bottom=Side(style='thin', color='FFD1D5DB')
)

alignments = [
    Alignment(horizontal='center', vertical='center'),  # Col A (Index)
    Alignment(horizontal='left', vertical='center'),    # Col B (Section)
    Alignment(horizontal='left', vertical='center'),    # Col C (Title)
    Alignment(horizontal='left', vertical='center'),    # Col D (URL Path)
    Alignment(horizontal='center', vertical='center'),  # Col E (Location)
    Alignment(horizontal='center', vertical='center'),  # Col F (Status)
    Alignment(horizontal='left', vertical='center', wrap_text=True), # Col G (Description)
    Alignment(horizontal='center', vertical='center')   # Col H (Timeline)
]

def setup_sheet_header(ws, title_text, subtitle_text):
    # Set row heights for headers
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 28.05
    
    # Title row
    ws.merge_cells('A1:H1')
    cell_a1 = ws['A1']
    cell_a1.value = title_text
    cell_a1.font = Font(name='Arial', size=14, bold=True, color='FF0F2D1F')
    cell_a1.alignment = Alignment(horizontal='left', vertical='center')
    
    # Subtitle row
    ws.merge_cells('A2:H2')
    cell_a2 = ws['A2']
    cell_a2.value = subtitle_text
    cell_a2.font = Font(name='Arial', size=9.5, italic=True, color='FF555555')
    cell_a2.alignment = Alignment(horizontal='left', vertical='center')
    
    # Legend row (Row 3)
    ws.cell(row=3, column=1).value = ""
    
    # Completed key
    ws.cell(row=3, column=2).value = "Completed"
    ws.cell(row=3, column=2).font = status_styles['Completed']['font']
    ws.cell(row=3, column=2).fill = status_styles['Completed']['fill']
    ws.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=2).border = thin_border
    
    # In Progress key
    ws.cell(row=3, column=3).value = "In Progress"
    ws.cell(row=3, column=3).font = status_styles['In Progress']['font']
    ws.cell(row=3, column=3).fill = status_styles['In Progress']['fill']
    ws.cell(row=3, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=3).border = thin_border
    
    # Not Started key
    ws.cell(row=3, column=4).value = "Not Started"
    ws.cell(row=3, column=4).font = status_styles['Not Started']['font']
    ws.cell(row=3, column=4).fill = status_styles['Not Started']['fill']
    ws.cell(row=3, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=4).border = thin_border
    
    for c in [1, 5, 6, 7, 8]:
        ws.cell(row=3, column=c).fill = fill_white
    
    # Table headers (Row 4)
    headers = ['#', 'Section', 'Page / Route Title', 'URL Path', 'Location', 'Status', 'Description / Remark', 'Timeline']
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = text
        cell.fill = fill_header
        cell.font = font_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col in [1, 5, 6, 8] else 'left', vertical='center')

def set_column_widths(ws):
    widths = {'A': 5.0, 'B': 22.0, 'C': 32.0, 'D': 42.0, 'E': 14.0, 'F': 13.0, 'G': 58.0, 'H': 18.0}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write_sheet_data(ws, all_rows):
    current_row = 5
    data_row_counter = 1
    for item in all_rows:
        is_header, data = item
        if is_header:
            ws.row_dimensions[current_row].height = 22.05
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_header
                cell.border = thin_border
                if col == 1:
                    cell.value = data
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            ws.row_dimensions[current_row].height = 34.05
            _, section, title, url_path, location, status, desc, timeline = data
            values = [data_row_counter, section, title, url_path, location, status, desc, timeline]
            data_row_counter += 1
            
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.value = values[col - 1]
                cell.border = thin_border
                cell.alignment = alignments[col - 1]
                cell.fill = fill_white
                
                if col == 3:
                    cell.font = font_title
                elif col == 6:
                    s_val = values[col - 1]
                    if s_val in status_styles:
                        cell.font = status_styles[s_val]['font']
                        cell.fill = status_styles[s_val]['fill']
                    else:
                        cell.font = font_regular
                else:
                    cell.font = font_regular
        current_row += 1

def main():
    print("Initializing Module Tracker spreadsheet reorganization...")
    
    # Define Core Modules Data (grouped by Feature Module)
    core_modules_data = [
        # ATTENDANCE
        (True, '  ATTENDANCE'),
        (False, (1, 'USER PORTAL', 'Attendance', '/attendance', 'User', 'Completed', 'Employee attendance dashboard — view clock-in/clock-out records, total hours, and attendance status by date.', 'Done')),
        (False, (3, 'ADMIN PANEL', 'HR – Attendance Scoping', '/admin/hr/attendance', 'Sub-page', 'Completed', 'Org-wide attendance dashboard — real-time clock-in roster, exceptions, geo-fencing, and log details.', 'Done')),
        (False, (4, 'ADMIN PANEL', 'HR – Attendance Records', '/admin/hr/attendance/records', 'Sub-page', 'Completed', 'Full historical search and export of employee attendance logs.', 'Done')),
        (False, (5, 'ADMIN PANEL', 'HR – Attendance Remote Access', '/admin/hr/attendance/remote-access', 'Sub-page', 'Not Started', 'Approve/revoke remote clock-in permissions for specific employees.', 'TBD')),
        (False, (6, 'ADMIN PANEL', 'HR – Site Locations', '/admin/hr/site-locations', 'Sub-page', 'Completed', 'Define geofences and IP restrictions for site-based employee clock-in.', 'Done')),
        (False, (7, 'DEPARTMENT CONSOLE', 'Dept HR – Attendance Daily', '/dept/[dept_id]/hr/attendance', 'Sub-page', 'Completed', 'Manager view of daily attendance clock-in roster for department employees.', 'Done')),
        (False, (8, 'DEPARTMENT CONSOLE', 'Dept HR – Attendance Logs', '/dept/[dept_id]/hr/attendance/records', 'Sub-page', 'Completed', 'Filtered search and export of department daily attendance history.', 'Done')),
        (False, (9, 'DEPARTMENT CONSOLE', 'Dept HR – Office Location', '/dept/[dept_id]/hr/office-location', 'Sub-page', 'Completed', 'Monitor geo-fenced boundaries for department employee attendance clock-in validation.', 'Done')),
        
        # LEAVE
        (True, '  LEAVE'),
        (False, (10, 'USER PORTAL', 'Leave', '/leave', 'User', 'In Progress', 'Employee leave management — view leave balance, apply for leave, track approval status, and view leave history.', '2 weeks')),
        (False, (11, 'USER PORTAL', 'Leave Request', '/leave/request', 'Sub-page', 'In Progress', 'Leave application form — employee selects leave type, date range, and reliever; triggers approval workflow.', '2 weeks')),
        (False, (12, 'ADMIN PANEL', 'HR – Leave Management', '/admin/hr/leave', 'Sub-page', 'In Progress', 'HR leave approval queue — review pending leave requests, check conflicts, and manage balances.', '2 weeks')),
        (False, (13, 'DEPARTMENT CONSOLE', 'Dept HR – Leave Approvals', '/dept/[dept_id]/hr/leave', 'Sub-page', 'In Progress', 'Supervisor review for department leave requests — reliever coverage and schedule conflict checking.', '2 weeks')),
        
        # PERFORMANCE MANAGEMENT (PMS)
        (True, '  PERFORMANCE MANAGEMENT (PMS)'),
        (False, (14, 'USER PORTAL', 'Goals (PMS)', '/goals', 'User', 'In Progress', 'Personal goal setting and tracking — employees set, update, and submit goals for approval.', '2 months')),
        (False, (15, 'USER PORTAL', 'Reviews / Performance', '/reviews', 'User', 'In Progress', 'Employee performance reviews — view submitted appraisals, ratings, and self-appraisals.', '2 months')),
        (False, (16, 'USER PORTAL', 'Performance Hub (PMS)', '/pms', 'User', 'In Progress', 'User-facing dashboard for goals, KPIs, reviews, and professional development.', '2 months')),
        (False, (17, 'USER PORTAL', 'PMS Goals', '/pms/goals', 'Sub-page', 'In Progress', 'Setup, update, and track personal goals and objectives.', '2 months')),
        (False, (18, 'USER PORTAL', 'PMS KPIs', '/pms/kpi', 'Sub-page', 'In Progress', 'View personal performance indicator targets and achievements.', '2 months')),
        (False, (19, 'USER PORTAL', 'PMS Peer Feedback', '/pms/peer-feedback', 'Sub-page', 'In Progress', 'Submit and review peer performance feedback.', '2 months')),
        (False, (20, 'USER PORTAL', 'PMS Reviews', '/pms/reviews', 'Sub-page', 'In Progress', 'Employee-facing performance evaluation reviews and self-appraisals.', '2 months')),
        (False, (21, 'USER PORTAL', 'PMS Attendance Link', '/pms/attendance', 'Sub-page', 'In Progress', 'View attendance contribution to overall performance review.', '2 months')),
        (False, (22, 'USER PORTAL', 'PMS Behaviour / Core Values', '/pms/behaviour', 'Sub-page', 'In Progress', 'Self-evaluation and review of core values alignment.', '2 months')),
        (False, (23, 'USER PORTAL', 'PMS CBT', '/pms/cbt', 'Sub-page', 'In Progress', 'Access training and assessment tests for performance review.', '1 week')),
        (False, (24, 'USER PORTAL', 'PMS Development Plans', '/pms/development-plans', 'Sub-page', 'In Progress', 'Track personal growth and professional development plan objectives.', '2 months')),
        (False, (25, 'ADMIN PANEL', 'HR – Performance Review (PMS)', '/admin/hr/pms', 'Sub-page', 'In Progress', 'Central panel for organizing performance review cycles, KPIs, core values, and appraisals.', '2 months')),
        (False, (26, 'ADMIN PANEL', 'HR – PMS Cycles', '/admin/hr/pms/cycles', 'Sub-page', 'In Progress', 'Create, schedule, and open/close performance review cycles.', '2 months')),
        (False, (27, 'ADMIN PANEL', 'HR – PMS Goals', '/admin/hr/pms/goals', 'Sub-page', 'In Progress', 'Review and approve employee-submitted goals and targets.', '2 months')),
        (False, (28, 'ADMIN PANEL', 'HR – PMS KPIs', '/admin/hr/pms/kpi', 'Sub-page', 'In Progress', 'Define global and department-specific KPIs and weights.', '2 months')),
        (False, (29, 'ADMIN PANEL', 'HR – PMS Reviews Appraisals', '/admin/hr/pms/reviews', 'Sub-page', 'In Progress', 'Access all employee appraisals, manage review assignments, and sign off reviews.', '2 months')),
        (False, (30, 'ADMIN PANEL', 'HR – PMS Competencies', '/admin/hr/pms/competencies', 'Sub-page', 'In Progress', 'Manage professional competency standards per department and role.', '2 months')),
        (False, (31, 'ADMIN PANEL', 'HR – PMS Behaviour / Values', '/admin/hr/pms/behaviour', 'Sub-page', 'In Progress', 'Setup appraisal criteria for core values and soft skills.', '2 months')),
        (False, (32, 'ADMIN PANEL', 'HR – PMS Attendance Score', '/admin/hr/pms/attendance', 'Sub-page', 'In Progress', 'Define attendance score contribution and rules for performance evaluations.', '2 months')),
        (False, (33, 'ADMIN PANEL', 'HR – PMS CBT Management', '/admin/hr/pms/cbt', 'Sub-page', 'In Progress', 'Manage computer-based tests, question banks, and review candidate scores.', '1 week')),
        (False, (34, 'ADMIN PANEL', 'HR – PMS Analytics Dashboard', '/admin/hr/pms/analytics', 'Sub-page', 'In Progress', 'Overall performance rating distribution, department scores, and growth analytics.', '2 months')),
        (False, (35, 'ADMIN PANEL', 'HR – PMS Development Plans', '/admin/hr/pms/development-plans', 'Sub-page', 'In Progress', 'Monitor employee professional development plans and training compliance.', '2 months')),
        (False, (36, 'DEPARTMENT CONSOLE', 'Dept HR – Performance (PMS)', '/dept/[dept_id]/hr/pms', 'Sub-page', 'In Progress', 'Departmental goals, appraisal review queue, and performance stats.', '2 months')),
        (False, (37, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Analytics', '/dept/[dept_id]/hr/pms/analytics', 'Sub-page', 'In Progress', 'Departmental performance review analytics, rating distributions, and targets.', '2 months')),
        (False, (38, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Goals Queue', '/dept/[dept_id]/hr/pms/goals', 'Sub-page', 'In Progress', 'Review and approve department employee goals and target configurations.', '2 months')),
        (False, (39, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Reviews Queue', '/dept/[dept_id]/hr/pms/reviews', 'Sub-page', 'In Progress', 'Supervisor performance review evaluation queue for department employees.', '2 months')),
        (False, (40, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Competencies', '/dept/[dept_id]/hr/pms/competencies', 'Sub-page', 'In Progress', 'View and manage professional competency standards for department roles.', '2 months')),
        (False, (41, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Behaviour', '/dept/[dept_id]/hr/pms/behaviour', 'Sub-page', 'In Progress', 'Assess department employees on core values and soft skills compliance.', '2 months')),
        (False, (42, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Attendance Score', '/dept/[dept_id]/hr/pms/attendance', 'Sub-page', 'In Progress', 'Monitor department employees daily attendance contribution to performance score.', '2 months')),
        (False, (43, 'DEPARTMENT CONSOLE', 'Dept HR – PMS CBT Results', '/dept/[dept_id]/hr/pms/cbt', 'Sub-page', 'In Progress', 'View department candidate scores for training and compliance tests.', '1 week')),
        (False, (44, 'DEPARTMENT CONSOLE', 'Dept HR – PMS Dev Plans', '/dept/[dept_id]/hr/pms/development-plans', 'Sub-page', 'In Progress', 'Track department employee growth and training plan compliance.', '2 months')),
        
        # TASKS
        (True, '  TASKS'),
        (False, (45, 'USER PORTAL', 'Tasks', '/tasks', 'Both', 'In Progress', 'Personal task board — employees view, filter, and update tasks assigned to them. Shared with Admin.', '2 months')),
        (False, (46, 'ADMIN PANEL', 'Tasks (Admin)', '/admin/tasks', 'Both', 'In Progress', 'Admin task management — view and manage all tasks across all employees and departments.', '2 months')),
        (False, (47, 'DEPARTMENT CONSOLE', 'Department Tasks', '/dept/[dept_id]/tasks', 'Department', 'In Progress', 'Team task board — create, assign, and track tasks for the department.', '2 months')),
        
        # FINANCE & PAYMENTS
        (True, '  FINANCE & PAYMENTS'),
        (False, (48, 'USER PORTAL', 'Payments', '/payments', 'User', 'In Progress', 'Employee payments ledger — displays salary, allowances, and reimbursement records with status.', 'unknown')),
        (False, (49, 'ADMIN PANEL', 'Finance Hub', '/admin/finance', 'Admin', 'In Progress', 'Finance overview dashboard — summary of invoices, bills, payments, and financial health indicators.', 'unknown')),
        (False, (50, 'ADMIN PANEL', 'Finance – Invoices', '/admin/finance/invoices', 'Sub-page', 'In Progress', 'Invoice management — create, view, and track outgoing invoices; status tracking.', 'unknown')),
        (False, (51, 'ADMIN PANEL', 'Finance – Bills', '/admin/finance/bills', 'Sub-page', 'In Progress', 'Vendor bills management — log, track, and approve bills; status tracking from draft to paid.', 'unknown')),
        (False, (52, 'ADMIN PANEL', 'Finance – Payments Ledger', '/admin/finance/payments', 'Sub-page', 'In Progress', 'Admin payments ledger — manage all employee or vendor payment records.', 'unknown')),
        (False, (53, 'ADMIN PANEL', 'Finance – Payments by Dept', '/admin/finance/payments/departments', 'Sub-page', 'In Progress', 'Department-level payment breakdown — view total payment obligations and disbursements.', 'unknown')),
        (False, (54, 'ADMIN PANEL', 'Finance – Reports', '/admin/finance/reports', 'Sub-page', 'In Progress', 'Financial reporting module — generates summary and detailed financial reports.', 'unknown')),
        (False, (55, 'ADMIN PANEL', 'Finance – Payments by Dept List', '/admin/payments/departments', 'Sub-page', 'In Progress', 'Consolidated view of payment schedules and salary totals per department.', 'unknown')),
        (False, (60, 'DEPARTMENT CONSOLE', 'Department Finance Hub', '/dept/[dept_id]/finance', 'Department', 'In Progress', 'Track department-specific invoices, bills, budgets, and payments.', 'unknown')),
        (False, (57, 'DEPARTMENT CONSOLE', 'Dept Finance – Bills', '/dept/[dept_id]/finance/bills', 'Sub-page', 'In Progress', 'Review, approve, and track vendor bills logged for the department.', 'unknown')),
        (False, (58, 'DEPARTMENT CONSOLE', 'Dept Finance – Invoices', '/dept/[dept_id]/finance/invoices', 'Sub-page', 'In Progress', 'Manage outgoing invoices generated by the department.', 'unknown')),
        (False, (59, 'DEPARTMENT CONSOLE', 'Dept Finance – Payments', '/dept/[dept_id]/finance/payments', 'Sub-page', 'In Progress', 'Department-level payment obligations and disbursement records.', 'unknown')),
        (False, (60, 'DEPARTMENT CONSOLE', 'Dept Finance – Reports', '/dept/[dept_id]/finance/reports', 'Sub-page', 'In Progress', 'Financial reporting module — generates departmental summaries and expense sheets.', 'unknown')),
        
        # DOCUMENTATION
        (True, '  DOCUMENTATION'),
        (False, (61, 'USER PORTAL', 'Documentation', '/documentation', 'Both', 'Completed', 'Employee document library — browse company policies, SOPs, and internal documents.', 'Done')),
        (False, (62, 'USER PORTAL', 'Department Documents', '/documentation/department', 'Sub-page', 'Completed', 'Department-specific document repository — filtered view of departmental files.', 'Done')),
        (False, (63, 'USER PORTAL', 'Internal Documents', '/documentation/internal', 'Sub-page', 'Completed', 'Internal company documents — company-wide policies, HR circulars, compliance documents.', 'Done')),
        (False, (64, 'ADMIN PANEL', 'Documentation (Admin)', '/admin/documentation', 'Both', 'Completed', 'Admin document management — upload, categorise, and publish company documents.', 'Done')),
        (False, (65, 'ADMIN PANEL', 'Documentation – Dept Docs', '/admin/documentation/department', 'Sub-page', 'Completed', 'Admin management of department-specific documents — upload, version, and control visibility.', 'Done')),
        (False, (66, 'ADMIN PANEL', 'Documentation – Internal Docs', '/admin/documentation/internal', 'Sub-page', 'Completed', 'Admin management of internal company-wide documents — HR policies, announcements.', 'Done')),
        (False, (67, 'DEPARTMENT CONSOLE', 'Department Document Library', '/dept/[dept_id]/documentation', 'Department', 'Completed', 'Manage files, policies, and standard operating procedures specific to the department.', 'Done')),
        (False, (68, 'DEPARTMENT CONSOLE', 'Dept Docs – Internal Docs', '/dept/[dept_id]/documentation/internal', 'Sub-page', 'Completed', 'Read-only access to company-wide internal circulars and policies.', 'Done')),
        (False, (69, 'DEPARTMENT CONSOLE', 'Dept Docs – Departmental', '/dept/[dept_id]/documentation/department', 'Sub-page', 'Completed', 'Upload, manage, and download documents and SOPs specific to the department.', 'Done')),
        
        # HELP DESK
        (True, '  HELP DESK'),
        (False, (70, 'USER PORTAL', 'Help Desk', '/help-desk', 'Both', 'In Progress', 'IT & general support ticket system — employees raise tickets for issues or requests.', '2 months')),
        (False, (71, 'ADMIN PANEL', 'Help Desk (Admin)', '/admin/help-desk', 'Both', 'In Progress', 'Admin support ticket management — full queue of raised tickets, staff assignment, and SLA tracking.', '2 months')),
        (False, (72, 'DEPARTMENT CONSOLE', 'Department Help Desk', '/dept/[dept_id]/help-desk', 'Department', 'In Progress', 'Handle support tickets assigned to or raised by the department.', '2 months')),
        
        # REPORTS & GENERAL MEETING
        (True, '  REPORTS & GENERAL MEETING'),
        (False, (73, 'USER PORTAL', 'Reports Hub', '/reports', 'Both', 'Completed', 'Employee-facing reports portal — access to weekly meeting reports, KSS, minutes, and trackers.', 'Done')),
        (False, (74, 'USER PORTAL', 'Weekly Reports', '/reports/weekly-reports', 'Sub-page', 'Completed', 'Lists auto-generated departmental weekly reports.', 'Done')),
        (False, (75, 'USER PORTAL', 'Knowledge Sharing Session', '/reports/kss', 'Sub-page', 'Completed', 'Knowledge Sharing Session documents portal — view and download KSS presentations.', 'Done')),
        (False, (76, 'USER PORTAL', 'Minutes of Meeting', '/reports/minutes-of-meeting', 'Sub-page', 'Completed', 'Access to uploaded Minutes of Meeting PDF documents — filterable by week and year.', 'Done')),
        (False, (77, 'USER PORTAL', 'Action Tracker', '/reports/action-tracker', 'Sub-page', 'Completed', 'Employee-facing action tracker — displays open and closed action items from general meetings.', 'Done')),
        (False, (78, 'USER PORTAL', 'General Meeting Reports Hub', '/reports/general-meeting', 'User', 'Completed', 'Main dashboard for general meeting outputs — weekly reports, KSS, minutes, and action items.', 'Done')),
        (False, (79, 'USER PORTAL', 'GM Action Tracker', '/reports/general-meeting/action-tracker', 'Sub-page', 'Completed', 'Detailed view of action items assigned during general meetings.', 'Done')),
        (False, (80, 'USER PORTAL', 'GM KSS', '/reports/general-meeting/kss', 'Sub-page', 'Completed', 'View and download Knowledge Sharing Session documents from general meetings.', 'Done')),
        (False, (81, 'USER PORTAL', 'GM Minutes', '/reports/general-meeting/minutes-of-meeting', 'Sub-page', 'Completed', 'Read and download official Minutes of Meeting PDFs from general meetings.', 'Done')),
        (False, (82, 'USER PORTAL', 'GM Weekly Reports', '/reports/general-meeting/weekly-reports', 'Sub-page', 'Completed', 'Access weekly reports from all departments submitted for general meetings.', 'Done')),
        (False, (83, 'USER PORTAL', 'Daily Activity Report', '/reports/daily-activity', 'User', 'Completed', 'Log and track daily work activity, submissions, and tasks performed.', 'Done')),
        (False, (84, 'ADMIN PANEL', 'Reports Admin Hub', '/admin/reports', 'Both', 'Completed', 'Admin reports management centre — upload, manage, and broadcast meeting reports, KSS, minutes.', 'Done')),
        (False, (85, 'ADMIN PANEL', 'Reports – Weekly Reports', '/admin/reports/weekly-reports', 'Sub-page', 'Completed', 'Admin-side weekly report management — view all departmental weekly reports.', 'Done')),
        (False, (86, 'ADMIN PANEL', 'Reports – KSS Upload', '/admin/reports/kss', 'Sub-page', 'Completed', 'Upload and manage Knowledge Sharing Session files.', 'Done')),
        (False, (87, 'ADMIN PANEL', 'Reports – Minutes Upload', '/admin/reports/minutes-of-meeting', 'Sub-page', 'Completed', 'Upload and manage Minutes of Meeting PDF documents.', 'Done')),
        (False, (88, 'ADMIN PANEL', 'Reports – Action Tracker', '/admin/reports/action-tracker', 'Sub-page', 'Completed', 'Admin action tracker management — create, update, and close action items.', 'Done')),
        (False, (89, 'ADMIN PANEL', 'Reports – Meeting Documents', '/admin/reports/meeting-documents', 'Sub-page', 'Completed', 'Master view of all meeting documents per week — KSS, minutes, and action point files.', 'Done')),
        (False, (90, 'ADMIN PANEL', 'Reports – Weekly Summary Mail', '/admin/reports/mail', 'Sub-page', 'Completed', 'Weekly meeting summary broadcast — send or schedule emails with PDF attachments.', 'Done')),
        (False, (91, 'ADMIN PANEL', 'GM Reports Hub Admin', '/admin/reports/general-meeting', 'Sub-page', 'Completed', 'Admin-side general meeting reports dashboard — weekly reports, KSS, minutes, and action items.', 'Done')),
        (False, (92, 'ADMIN PANEL', 'GM Action Tracker Admin', '/admin/reports/general-meeting/action-tracker', 'Sub-page', 'Completed', 'Admin action items manager for general meetings.', 'Done')),
        (False, (93, 'ADMIN PANEL', 'GM KSS Admin', '/admin/reports/general-meeting/kss', 'Sub-page', 'Completed', 'Upload and catalog Knowledge Sharing Session presentations for general meetings.', 'Done')),
        (False, (94, 'ADMIN PANEL', 'GM Minutes Admin', '/admin/reports/general-meeting/minutes-of-meeting', 'Sub-page', 'Completed', 'Upload and publish official general meeting minutes.', 'Done')),
        (False, (95, 'ADMIN PANEL', 'GM Weekly Reports Admin', '/admin/reports/general-meeting/weekly-reports', 'Sub-page', 'Completed', 'Review and merge weekly reports from all departments.', 'Done')),
        (False, (96, 'DEPARTMENT CONSOLE', 'Department Reports Hub', '/dept/[dept_id]/reports', 'Department', 'Completed', 'Manage department weekly meeting reports, minutes, and action trackers.', 'Done')),
        (False, (97, 'DEPARTMENT CONSOLE', 'Dept Reports – Weekly Reports', '/dept/[dept_id]/reports/weekly', 'Sub-page', 'Completed', 'Submit and edit department weekly reports for general meetings.', 'Done')),
        (False, (98, 'DEPARTMENT CONSOLE', 'Dept HR – Daily Activity Logs', '/dept/[dept_id]/hr/reports/daily-activity', 'Sub-page', 'Completed', 'Review daily work logs submitted by employees within the department.', 'Done')),
        
        # FEEDBACK
        (True, '  FEEDBACK'),
        (False, (99, 'USER PORTAL', 'Feedback', '/feedback', 'Both', 'Completed', 'Anonymous or named feedback submission — employees submit suggestions, concerns, or recognition.', 'Done')),
        (False, (100, 'ADMIN PANEL', 'Feedback (Admin)', '/admin/feedback', 'Both', 'Completed', 'Admin feedback inbox — review employee feedback submissions; respond and resolve.', 'Done')),
        (False, (101, 'DEPARTMENT CONSOLE', 'Department Feedback', '/dept/[dept_id]/feedback', 'Department', 'Completed', 'Review feedback from team members within the department.', 'Done')),
        
        # ASSETS
        (True, '  ASSETS'),
        (False, (102, 'USER PORTAL', 'Assets', '/assets', 'Both', 'Completed', 'Employee view of company assets — lists assets assigned to the individual or department.', 'Done')),
        (False, (103, 'ADMIN PANEL', 'Assets (Admin)', '/admin/assets', 'Both', 'Completed', 'Admin asset management — full company asset register with assignment, status, and logs.', 'Done')),
        (False, (104, 'ADMIN PANEL', 'Assets – Issue Tracking', '/admin/assets/issues', 'Sub-page', 'Completed', 'Track reported asset issues and faults — log damage reports and manage repairs.', 'Done')),
        (False, (105, 'DEPARTMENT CONSOLE', 'Department Assets', '/dept/[dept_id]/assets', 'Department', 'Completed', 'View and request company assets assigned to the department.', 'Done')),
        
        # CORRESPONDENCE & COMMUNICATIONS
        (True, '  CORRESPONDENCE & COMMUNICATIONS'),
        (False, (106, 'USER PORTAL', 'Correspondence / Reference Letters', '/correspondence', 'User', 'Completed', 'Employee-facing correspondence portal — manage and generate official letters.', 'Done')),
        (False, (107, 'ADMIN PANEL', 'Correspondence Administration', '/admin/correspondence', 'Admin', 'Completed', 'Admin correspondence dashboard — generate and approve reference/intro letters.', 'Done')),
        (False, (108, 'ADMIN PANEL', 'Communications – Broadcast', '/admin/communications/broadcast', 'Sub-page', 'Completed', 'Email/notification broadcast centre — compose and send targeted messages.', 'Done')),
        (False, (109, 'ADMIN PANEL', 'Communications – Meetings', '/admin/communications/meetings', 'Sub-page', 'Completed', 'Meeting communications management — view meeting schedule and coord invites.', 'Done')),
        (False, (110, 'ADMIN PANEL', 'Communications – Meeting Mail', '/admin/communications/meetings/mail', 'Sub-page', 'Completed', 'Send formal meeting invitations and follow-ups.', 'Done')),
        (False, (111, 'ADMIN PANEL', 'Communications – Reminders', '/admin/communications/meetings/reminders', 'Sub-page', 'Completed', 'Schedule automated meeting reminder emails.', 'Done')),
        (False, (112, 'DEPARTMENT CONSOLE', 'Department Correspondence', '/dept/[dept_id]/correspondence', 'Department', 'Completed', 'Handle department-specific correspondence and official letter templates.', 'Done')),
        (False, (113, 'DEPARTMENT CONSOLE', 'Department Communications', '/dept/[dept_id]/communications', 'Department', 'Completed', 'Send internal communications and announcements to department members.', 'Done')),
        (False, (114, 'DEPARTMENT CONSOLE', 'Dept Communications – Broadcast', '/dept/[dept_id]/communications/broadcast', 'Sub-page', 'Completed', 'Compose and broadcast emails/notifications to department members.', 'Done')),
        
        # EMPLOYEES & DIRECTORY
        (True, '  EMPLOYEES & DIRECTORY'),
        (False, (115, 'USER PORTAL', 'Profile / Home', '/profile', 'User', 'Completed', 'Employee personal dashboard — shows profile info, department, role, and quick links.', 'Done')),
        (False, (116, 'USER PORTAL', 'Employee Directory', '/directory', 'User', 'Completed', 'Interactive directory showing all active employee profiles, departments, and roles.', 'Done')),
        (False, (117, 'ADMIN PANEL', 'HR Management Hub', '/admin/hr', 'Admin', 'Completed', 'Central HR administration panel — overview of employees, leaves, and departments.', 'Done')),
        (False, (118, 'ADMIN PANEL', 'HR – Employees List', '/admin/hr/employees', 'Sub-page', 'Completed', 'Full employee directory — HR can view, edit, or deactivate profiles.', 'Done')),
        (False, (119, 'ADMIN PANEL', 'HR – Departments', '/admin/hr/departments', 'Sub-page', 'Completed', 'Department management — create, edit, and manage company departments.', 'Done')),
        (False, (121, 'ADMIN PANEL', 'HR – Office Location', '/admin/hr/office-location', 'Sub-page', 'Completed', 'Manage registered office locations — define geo-fence boundaries for clock-in.', 'Done')),
        (False, (122, 'DEPARTMENT CONSOLE', 'Department HR Hub', '/dept/[dept_id]/hr', 'Department', 'Completed', 'Manager view of department employee profiles, roles, and status.', 'Done')),
        (False, (123, 'DEPARTMENT CONSOLE', 'Dept HR – Employees', '/dept/[dept_id]/hr/employees', 'Sub-page', 'Completed', 'View active profiles, contact info, and roles of department employees.', 'Done')),
        (False, (124, 'DEPARTMENT CONSOLE', 'Dept HR – Departments', '/dept/[dept_id]/hr/departments', 'Sub-page', 'Completed', 'View department hierarchy, structures, and employee counts.', 'Done'))
    ]
    
    # Define Miscellaneous Sheet Data (standalone pages)
    misc_sheet_data = [
        # AUTH PAGES
        (True, '  AUTH PAGES'),
        (False, (1, 'AUTH PAGES', 'Login', '/auth/login', 'Auth', 'Completed', 'Employee login page — email and password authentication with session management.', 'Done')),
        (False, (2, 'AUTH PAGES', 'Sign Up', '/auth/sign-up', 'Auth', 'Completed', 'New account registration — used during employee onboarding via invite link.', 'Done')),
        (False, (3, 'AUTH PAGES', 'Sign Up Success', '/auth/sign-up-success', 'Auth', 'Completed', 'Post-registration confirmation page.', 'Done')),
        (False, (4, 'AUTH PAGES', 'Forgot Password', '/auth/forgot-password', 'Auth', 'Completed', 'Password reset request.', 'Done')),
        (False, (5, 'AUTH PAGES', 'Reset Password', '/auth/reset-password', 'Auth', 'Completed', 'Password reset form.', 'Done')),
        (False, (6, 'AUTH PAGES', 'Set Password', '/auth/set-password', 'Auth', 'Completed', 'Initial password setup page for newly invited employees.', 'Done')),
        (False, (7, 'AUTH PAGES', 'Setup Account / Onboarding', '/auth/setup-account', 'Auth', 'Completed', 'New employee account setup wizard.', 'Done')),
        (False, (8, 'AUTH PAGES', 'Auth Error Page', '/auth/error', 'Auth', 'Completed', 'Auth error boundary page.', 'Done')),
        
        # DEVELOPER TOOLS
        (True, '  DEVELOPER TOOLS'),
        (False, (9, 'DEVELOPER TOOLS', 'Dev Dashboard', '/admin/dev', 'Admin (Dev)', 'Completed', 'Developer-only section overview — quick links to all diagnostics.', 'Done')),
        (False, (10, 'DEVELOPER TOOLS', 'Login Logs', '/admin/dev/login-logs', 'Admin (Dev)', 'Completed', 'System login activity log — sign-in events with timestamp, IP, and success status.', 'Done')),
        (False, (11, 'DEVELOPER TOOLS', 'Role Escalations', '/admin/dev/role-escalations', 'Admin (Dev)', 'Completed', 'Tracks temporary role escalation events.', 'Done')),
        (False, (12, 'DEVELOPER TOOLS', 'Security Events', '/admin/dev/security-events', 'Admin (Dev)', 'Completed', 'Security incident log — records suspicious actions and failed auth.', 'Done')),
        (False, (13, 'DEVELOPER TOOLS', 'Tests', '/admin/dev/tests', 'Admin (Dev)', 'Completed', 'Built-in integration test runner — execute E2E tests.', 'Done')),
        (False, (14, 'DEVELOPER TOOLS', 'UI Error Monitor', '/admin/dev/ui-errors', 'Admin (Dev)', 'Completed', 'Frontend error tracking — logs JS and React errors.', 'Done')),
        (False, (15, 'DEVELOPER TOOLS', 'Impersonation Tool', '/admin/dev/impersonation', 'Admin (Dev)', 'Completed', 'Developer diagnostic tool for impersonating employee accounts.', 'Done')),
        (False, (16, 'DEVELOPER TOOLS', 'Acobot Diagnostics', '/admin/dev/acobot', 'Admin (Dev)', 'Completed', 'AI diagnostics — monitor context, prompts, and responses.', 'Done')),
        
        # SYSTEM SETTINGS
        (True, '  SYSTEM SETTINGS'),
        (False, (17, 'ADMIN PANEL', 'Settings', '/admin/settings', 'Admin', 'Completed', 'Top-level admin settings hub — manage company config and system maintenance.', 'Done')),
        (False, (18, 'ADMIN PANEL', 'Settings – Company', '/admin/settings/company', 'Sub-page', 'Completed', 'Company profile settings — update name, address, and branding.', 'Done')),
        (False, (19, 'ADMIN PANEL', 'Settings – Mail', '/admin/settings/mail', 'Sub-page', 'Completed', 'Email system configuration — manage mail provider settings and test.', 'Done')),
        (False, (20, 'ADMIN PANEL', 'Settings – Maintenance', '/admin/settings/maintenance', 'Sub-page', 'Completed', 'System maintenance mode toggle and scheduling.', 'Done')),
        (False, (21, 'ADMIN PANEL', 'Settings – Roles & Permissions', '/admin/settings/roles', 'Sub-page', 'Completed', 'Role-based access control (RBAC) management — define roles and scopes.', 'Done')),
        (False, (22, 'ADMIN PANEL', 'Settings – User Management', '/admin/settings/users', 'Sub-page', 'Completed', 'Manage system user accounts — view, invite, or reset passwords.', 'Done')),
        
        # INVENTORY & PURCHASING
        (True, '  INVENTORY & PURCHASING'),
        (False, (23, 'ADMIN PANEL', 'Inventory Hub', '/admin/inventory', 'Admin', 'Completed', 'Admin inventory management — overview of products, stock levels, and warehouse zones.', 'Done')),
        (False, (24, 'ADMIN PANEL', 'Inventory – Products', '/admin/inventory/products', 'Sub-page', 'Completed', 'Full product catalogue — manage SKU, category, and stock.', 'Done')),
        (False, (25, 'ADMIN PANEL', 'Inventory – Categories', '/admin/inventory/categories', 'Sub-page', 'Completed', 'Product category management — create and organize categories.', 'Done')),
        (False, (26, 'ADMIN PANEL', 'Inventory – Warehouses', '/admin/inventory/warehouses', 'Sub-page', 'Completed', 'Warehouse location management — register storage zones.', 'Done')),
        (False, (27, 'ADMIN PANEL', 'Inventory – Stock Movements', '/admin/inventory/movements', 'Sub-page', 'Completed', 'Stock movement log — tracks all stock adjustments.', 'Done')),
        (False, (28, 'ADMIN PANEL', 'Purchasing Hub', '/admin/purchasing', 'Admin', 'Completed', 'Procurement dashboard — supplier performance, order queues.', 'Done')),
        (False, (29, 'ADMIN PANEL', 'Purchasing – Orders', '/admin/purchasing/orders', 'Sub-page', 'Completed', 'Purchase order management — creation, tracking, and receipts.', 'Done')),
        (False, (30, 'ADMIN PANEL', 'Purchasing – Suppliers', '/admin/purchasing/suppliers', 'Sub-page', 'Completed', 'Supplier directory — contact info, terms, and supplier ratings.', 'Done')),
        (False, (31, 'ADMIN PANEL', 'Purchasing – Receipts', '/admin/purchasing/receipts', 'Sub-page', 'Completed', 'Goods receipt log — confirm deliveries against POs.', 'Done')),
        
        # STANDALONE PORTAL TOOLS
        (True, '  STANDALONE PORTAL TOOLS'),
        (False, (32, 'USER PORTAL', 'Tools Hub', '/tools', 'Both', 'Completed', 'Self-service tool portal — signature generator, watermark, etc.', 'Done')),
        (False, (33, 'USER PORTAL', 'Email Signature Generator', '/tools/signature', 'Sub-page', 'Completed', 'Generates branded ACOB HTML signature using profile data.', 'Done')),
        (False, (34, 'USER PORTAL', 'Document Watermark Tool', '/tools/watermark', 'Sub-page', 'Completed', 'Applies watermarks to uploaded PDF documents.', 'Done')),
        (False, (35, 'USER PORTAL', 'Anniversary Signature Tool', '/tools/signature-anniversary', 'Sub-page', 'Completed', 'Generates custom anniversary-themed email signatures for employees.', 'Done'))
    ]
    
    # Load workbook
    wb = openpyxl.load_workbook(WORKSPACE_XLSX)
    
    # --- Rebuild Sheet 1: Module Tracker ---
    print("Rebuilding main sheet 'Module Tracker'...")
    ws = wb['Module Tracker']
    ws.title = 'Module Tracker'
    
    # Clear old merges and rows
    initial_merged = list(ws.merged_cells.ranges)
    ranges_to_keep = [r for r in initial_merged if r.min_row <= 2]
    ws.merged_cells.ranges = ranges_to_keep
    
    max_row = ws.max_row
    if max_row >= 5:
        ws.delete_rows(5, max_row - 4)
        
    setup_sheet_header(ws, 'ACOB LIGHTING TECHNOLOGY LIMITED — Matrix System Module Development Tracker', 
                       'Tracks core system modules by matching equivalents across the User Portal, Admin Panel, and Department Console.')
    set_column_widths(ws)
    write_sheet_data(ws, core_modules_data)
    
    # --- Rebuild/Create Sheet 2: Miscellaneous ---
    print("Rebuilding second sheet 'Miscellaneous'...")
    if 'Miscellaneous' in wb.sheetnames:
        ws_misc = wb['Miscellaneous']
        # Clear it completely
        ws_misc.merged_cells.ranges = []
        ws_misc.delete_rows(1, ws_misc.max_row + 1)
    else:
        ws_misc = wb.create_sheet(title='Miscellaneous')
        
    setup_sheet_header(ws_misc, 'ACOB LIGHTING TECHNOLOGY LIMITED — Miscellaneous Standalone Routes', 
                       'Tracks standalone, authentication, configuration, and utility pages not tied to core department-spanning modules.')
    set_column_widths(ws_misc)
    write_sheet_data(ws_misc, misc_sheet_data)
    
    # Save workspace copy
    print(f"Saving workspace file to {WORKSPACE_XLSX}...")
    wb.save(WORKSPACE_XLSX)
    
    # Copy to downloads
    print(f"Copying updated file to downloads {DOWNLOADS_XLSX}...")
    try:
        shutil.copy(WORKSPACE_XLSX, DOWNLOADS_XLSX)
        print("Success! Reorganization completed successfully.")
    except Exception as e:
        print(f"Error copying to downloads: {e}")

if __name__ == '__main__':
    main()
