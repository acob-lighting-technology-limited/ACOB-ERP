import openpyxl
import datetime
import dotenv
import psycopg2
import re

env = dotenv.dotenv_values('.env.local')
db_url = env.get('SUPABASE_POSTGRES_URL_NON_POOLING') or env.get('SUPABASE_POSTGRES_URL')
conn = psycopg2.connect(db_url)
cur = conn.cursor()

cur.execute('SELECT id, full_name, first_name, last_name FROM profiles;')
db_profiles = cur.fetchall()
prof_by_id = {p[0]: p[1] for p in db_profiles}

def normalize(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s)).strip().lower()

MANUAL_NAME_MAP = {
    'shirley': 'Shirley Jackreece',
    'john': 'John Dangana',
    'susan eze': 'Susan Eze',
    'emmanuel': 'Emmanuel Ibanga',
    'vanessa': 'Vanessa Lawrence-Ukaegbu',
    'thomas': 'Thomas Olobo',
    'philip': 'Philip Yakubu',
    'busayo': 'Busayo Kadiri',
    'edward': 'Edward Atoshi',
    'andrew': 'Andrew Inegbedion',
    'kennedy': 'Kennedy Odenigbo',
    'tansi': 'Tansi',
    'dennis': 'Dennis Innug',
    'lawrence': 'Lawrence Adukwu',
    'favour': 'Favour Onuoha-Eke',
    'jessica': 'Jessica Egeonu',
    'bawa': 'Alhamdu Bawa',
    'samila': 'Agya Samila',
    'elijah': 'Elijah Isah',
    'tunde': 'Tunde Ajayi',
    'patrick': 'Patrick Prosper',
    'arc emmanuel': 'Emmanuel Chinedu',
    'tochucku': 'Tochukwu Nnadozie',
    'samad': 'Abdulsamad Danmusa',
    'abiodun': 'Sefui Abiodun',
    'chibuikim': 'Chibuikem Ilonze',
    'victor': 'Victor Onyeka',
    'usman': 'Usman Haruna',
    'habeeb': 'Habeeb',
    'suraj': 'Surajo Idris',
    'caleb': 'Caleb Obiechina',
    'daniel': 'Daniel Osa-Egonwa',
    'anointing': 'Anointed Emoghene',
    'oghenerune': 'Oghenerune Orhorhomuke',
    'benson': 'Benson Ademoye',
    'sommy': 'Sommy',
    'ayoola': 'Peter Ayoola',
    'onyeka': 'Onyeka Eze',
    'oyneka': 'Onyeka Eze',
    'khleo': 'Kleopatra Aiyede',
    'abdul': 'Abdulmalik Abdulkarim',
    'haron': 'Ode Haron',
    'peace': 'Peace Dogara',
    'new daniel': 'Istifanus Daniel',
    'seun': 'Oluwaseun Awotona',
    'tobi': 'Oluwatobi Oladele'
}

def find_profile(name):
    norm = normalize(name)
    if not norm or norm in ['total', 'names', 'date', 's/n']: return None
    if norm in MANUAL_NAME_MAP:
        mapped_name = MANUAL_NAME_MAP[norm]
        for p in db_profiles:
            if normalize(p[1]) == normalize(mapped_name):
                return p
    for p in db_profiles:
        if normalize(p[1]) == norm or normalize(p[2]) == norm or normalize(p[3]) == norm:
            return p
    return None

# Parse File 1 (ACOB PETTY CASH BOOKS.xlsx)
wb_old = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\ACOB PETTY CASH BOOKS.xlsx', data_only=True)
ws_lakita = wb_old['LAKITA FOOD 2026']
ws_ded = wb_old['FOOD DEDUCTION 2026']

headers_l = [ws_lakita.cell(1, c).value for c in range(1, ws_lakita.max_column + 1)]
headers_d = [ws_ded.cell(1, c).value for c in range(1, ws_ded.max_column + 1)]

emp_map_old = {}
for c in range(3, 51):
    h = headers_l[c-1] or headers_d[c-1]
    p = find_profile(h)
    if p:
        emp_map_old[c] = p

def parse_date_old(val):
    dt = None
    if isinstance(val, (datetime.datetime, datetime.date)): dt = val.date() if isinstance(val, datetime.datetime) else val
    elif isinstance(val, str):
        try: dt = datetime.datetime.strptime(val.strip(), '%Y-%m-%d').date()
        except: dt = None
    if dt:
        if dt.year == 2025: dt = datetime.date(2026, dt.month, dt.day)
        if dt == datetime.date(2026, 1, 31): dt = datetime.date(2026, 1, 30)
    return dt

old_file_entries = {}
for ws in [ws_lakita, ws_ded]:
    for r in range(1, ws.max_row + 1):
        dt = parse_date_old(ws.cell(r, 2).value)
        if dt:
            for c in range(3, 51):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v > 0:
                    if c in emp_map_old:
                        p = emp_map_old[c]
                        old_file_entries[(dt, p[0])] = (p[1], v)

# Fetch Updated Database logs
cur.execute("SELECT date, user_id FROM attendance_lunch_log WHERE date >= '2026-01-01' AND date <= '2026-12-31';")
updated_db_logs = set(cur.fetchall())

print("=================================================================")
print("  MONTH-BY-MONTH DISCREPANCY DETAILED SUMMARY (JAN - JUN 2026)")
print("=================================================================\n")

months = range(1, 7)
month_names = ['January', 'February', 'March', 'April', 'May', 'June']

total_f1_all = 0
total_db_all = 0

for m, m_name in zip(months, month_names):
    f1_m = set(k for k in old_file_entries.keys() if k[0].month == m)
    db_m = set(k for k in updated_db_logs if k[0].month == m)
    
    total_f1_all += len(f1_m)
    total_db_all += len(db_m)
    
    in_db_not_f1 = db_m - f1_m
    in_f1_not_db = f1_m - db_m
    
    print(f"=== {m_name.upper()} 2026 ===")
    print(f"File 1 Total: {len(f1_m)} meals | Database Total: {len(db_m)} meals | Net Difference: {len(db_m) - len(f1_m):+d}")
    print(f"  - Meals in Database NOT in File 1: {len(in_db_not_f1)}")
    print(f"  - Meals in File 1 NOT in Database: {len(in_f1_not_db)}")
    
    if in_db_not_f1:
        print("\n  [Sample Meals in DB but Missing in File 1]:")
        for dt, uid in sorted(in_db_not_f1)[:10]:
            name = prof_by_id.get(uid, uid)
            print(f"    - Date {dt} | {name}")
        if len(in_db_not_f1) > 10:
            print(f"    ... and {len(in_db_not_f1) - 10} more!")

    if in_f1_not_db:
        print("\n  [Sample Meals in File 1 but Missing in DB]:")
        for dt, uid in sorted(in_f1_not_db)[:10]:
            name = prof_by_id.get(uid, uid)
            print(f"    - Date {dt} | {name}")
        if len(in_f1_not_db) > 10:
            print(f"    ... and {len(in_f1_not_db) - 10} more!")
            
    print("\n" + "-" * 65 + "\n")

print(f"OVERALL JAN - JUN TOTALS:")
print(f"File 1 Total (Jan-Jun): {total_f1_all}")
print(f"Database Total (Jan-Jun): {total_db_all}")
print(f"Net Difference (Jan-Jun): {total_db_all - total_f1_all:+d}")

conn.close()
