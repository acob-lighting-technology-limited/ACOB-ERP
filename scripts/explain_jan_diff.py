import openpyxl
import datetime
import dotenv
import psycopg2
import re

env = dotenv.dotenv_values('.env.local')
db_url = env.get('SUPABASE_POSTGRES_URL_NON_POOLING') or env.get('SUPABASE_POSTGRES_URL')
conn = psycopg2.connect(db_url)
cur = conn.cursor()

cur.execute('SELECT id, full_name, first_name, last_name FROM profiles;')
db_profiles = cur.fetchall()
prof_by_id = {p[0]: p[1] for p in db_profiles}

def normalize(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s)).strip().lower()

MANUAL_NAME_MAP = {
    'shirley': 'Shirley Jackreece',
    'john': 'John Dangana',
    'susan eze': 'Susan Eze',
    'emmanuel': 'Emmanuel Ibanga',
    'vanessa': 'Vanessa Lawrence-Ukaegbu',
    'thomas': 'Thomas Olobo',
    'philip': 'Philip Yakubu',
    'busayo': 'Busayo Kadiri',
    'edward': 'Edward Atoshi',
    'andrew': 'Andrew Inegbedion',
    'kennedy': 'Kennedy Odenigbo',
    'tansi': 'Tansi',
    'dennis': 'Dennis Innug',
    'lawrence': 'Lawrence Adukwu',
    'favour': 'Favour Onuoha-Eke',
    'jessica': 'Jessica Egeonu',
    'bawa': 'Alhamdu Bawa',
    'samila': 'Agya Samila',
    'elijah': 'Elijah Isah',
    'tunde': 'Tunde Ajayi',
    'patrick': 'Patrick Prosper',
    'arc emmanuel': 'Emmanuel Chinedu',
    'tochucku': 'Tochukwu Nnadozie',
    'samad': 'Abdulsamad Danmusa',
    'abiodun': 'Sefui Abiodun',
    'chibuikim': 'Chibuikem Ilonze',
    'victor': 'Victor Onyeka',
    'usman': 'Usman Haruna',
    'habeeb': 'Habeeb',
    'suraj': 'Surajo Idris',
    'caleb': 'Caleb Obiechina',
    'daniel': 'Daniel Osa-Egonwa',
    'anointing': 'Anointed Emoghene',
    'oghenerune': 'Oghenerune Orhorhomuke',
    'benson': 'Benson Ademoye',
    'sommy': 'Sommy',
    'ayoola': 'Peter Ayoola',
    'onyeka': 'Onyeka Eze',
    'oyneka': 'Onyeka Eze',
    'khleo': 'Kleopatra Aiyede',
    'abdul': 'Abdulmalik Abdulkarim',
    'haron': 'Ode Haron',
    'peace': 'Peace Dogara',
    'new daniel': 'Istifanus Daniel',
    'seun': 'Oluwaseun Awotona',
    'tobi': 'Oluwatobi Oladele'
}

def find_profile(name):
    norm = normalize(name)
    if not norm or norm in ['total', 'names', 'date', 's/n']: return None
    if norm in MANUAL_NAME_MAP:
        mapped_name = MANUAL_NAME_MAP[norm]
        for p in db_profiles:
            if normalize(p[1]) == normalize(mapped_name):
                return p
    for p in db_profiles:
        if normalize(p[1]) == norm or normalize(p[2]) == norm or normalize(p[3]) == norm:
            return p
    return None

# Parse File 1 Jan entries
wb_old = openpyxl.load_workbook(r'C:\Users\IT_COMMS\Downloads\ACOB PETTY CASH BOOKS.xlsx', data_only=True)
ws_lakita = wb_old['LAKITA FOOD 2026']
ws_ded = wb_old['FOOD DEDUCTION 2026']

headers_l = [ws_lakita.cell(1, c).value for c in range(1, ws_lakita.max_column + 1)]
headers_d = [ws_ded.cell(1, c).value for c in range(1, ws_ded.max_column + 1)]

emp_map_old = {}
for c in range(3, 51):
    h = headers_l[c-1] or headers_d[c-1]
    p = find_profile(h)
    if p:
        emp_map_old[c] = p

def parse_date_old(val):
    dt = None
    if isinstance(val, (datetime.datetime, datetime.date)): dt = val.date() if isinstance(val, datetime.datetime) else val
    elif isinstance(val, str):
        try: dt = datetime.datetime.strptime(val.strip(), '%Y-%m-%d').date()
        except: dt = None
    if dt:
        if dt.year == 2025: dt = datetime.date(2026, dt.month, dt.day)
        if dt == datetime.date(2026, 1, 31): dt = datetime.date(2026, 1, 30)
    return dt

f1_jan = set()
for ws in [ws_lakita, ws_ded]:
    for r in range(1, 20):
        dt = parse_date_old(ws.cell(r, 2).value)
        if dt and dt.month == 1 and dt.year == 2026:
            for c in range(3, 51):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v > 0:
                    if c in emp_map_old:
                        f1_jan.add((dt, emp_map_old[c][0]))

# Parse DB Jan entries
cur.execute("SELECT date, user_id FROM attendance_lunch_log WHERE date >= '2026-01-01' AND date <= '2026-01-31';")
db_jan = set(cur.fetchall())

print(f"File 1 Jan total mapped meals: {len(f1_jan)}")
print(f"Updated DB Jan total meals: {len(db_jan)}")
print(f"Net Difference: {len(db_jan) - len(f1_jan)}")

in_db_not_f1 = db_jan - f1_jan
in_f1_not_db = f1_jan - db_jan

print(f"\n1. In DB but NOT in File 1 ({len(in_db_not_f1)} entries):")
for dt, uid in sorted(in_db_not_f1):
    name = prof_by_id.get(uid, uid)
    print(f"  - Date: {dt} | Employee: {name}")

print(f"\n2. In File 1 but NOT in DB ({len(in_f1_not_db)} entries):")
for dt, uid in sorted(in_f1_not_db):
    name = prof_by_id.get(uid, uid)
    print(f"  - Date: {dt} | Employee: {name}")

conn.close()
